"""PERV Website backend.

Single-file Flask application that:
- preprocesses sequence/* (xlsx, bed, fasta) into data/*.json on startup
- exposes JSON APIs for the two website sections
- serves Jinja2 templates and static assets

Run:
    pip install -r requirements.txt
    flask --app app run --host 0.0.0.0 --port 5000
or for production:
    gunicorn -w 2 -b 0.0.0.0:5000 app:app
"""

from __future__ import annotations

import io
import json
import os
import re
import tempfile
import time
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from flask import (
    Flask,
    Response,
    abort,
    jsonify,
    render_template,
    request,
    send_file,
    send_from_directory,
)

import genome as genome_lib


BASE_DIR = Path(__file__).resolve().parent
SEQ_DIR = BASE_DIR / "sequence"
DATA_DIR = BASE_DIR / "data"
GENOME_REF_DIR = BASE_DIR / "genome.ref.guochao"
DATA_DIR.mkdir(exist_ok=True)

XLSX_PATH = SEQ_DIR / "1165.intact.PERV.infomation.xlsx"
GENOME_INFO_PATH = SEQ_DIR / "genome.information.xlsx"
ALL_FA_PATH = SEQ_DIR / "my.final.fa"
PASS_FA_PATH = SEQ_DIR / "pass.139.fa"
ORF_BED_PATH = SEQ_DIR / "ORF.combine.HTML.bed"
DOMAIN_BED_PATH = SEQ_DIR / "domin.combine.HTML.bed"

META_JSON = DATA_DIR / "meta_1165.json"
ORF_JSON = DATA_DIR / "orf_index.json"
DOMAIN_JSON = DATA_DIR / "domain_index.json"
SEQ_OFFSETS_JSON = DATA_DIR / "seq_offsets.json"

GENOME_FASTA = GENOME_REF_DIR / "Sus_scrofa.Sscrofa11.1.dna.toplevel.fa"
GENOME_GTF = GENOME_REF_DIR / "Sus_scrofa.Sscrofa11.1.108.gtf"
GENOME_FAI = DATA_DIR / "genome.fa.fai"
GENOME_DB = DATA_DIR / "gtf.sqlite"
GENOME_BED = DATA_DIR / "genome.bed"
GENOME_GENES_BED = DATA_DIR / "genome.genes.bed"

MULTIOMICS_DIR = BASE_DIR / "new.Multi-omics"
MULTIOMICS_META = MULTIOMICS_DIR / "all.sample.info"
MULTIOMICS_REPRESENT = MULTIOMICS_DIR / "represent.sample.info"

PERV_REGION_FILE = BASE_DIR / "Homologous" / "RF.intact.region"
PERV_BED = DATA_DIR / "perv.bed"

HOMOLOGOUS_XLSX = BASE_DIR / "Homologous" / "final.Statistics.table.xlsx"
HOMOLOGOUS_SEQ_BED = DATA_DIR / "homologous_seq.bed"
HOMOLOGOUS_LOCUS_BED = DATA_DIR / "homologous_locus.bed"

DOWNLOAD_WHITELIST = {
    "my.final.fa",
    "pass.139.fa",
    "1165.intact.PERV.infomation.xlsx",
    "ORF.combine.HTML.bed",
    "domin.combine.HTML.bed",
}

# Maximum DNA region (in bp) returned by the genome API.
MAX_DNA_REGION = 1_000_000
# Maximum features returned by the genome features API.
MAX_FEATURES = 5000


# ---------------------------------------------------------------------------
# Sequence utilities: standard codon table, translation, reverse complement
# ---------------------------------------------------------------------------

STANDARD_CODON_TABLE: dict[str, str] = {
    "TTT": "F", "TTC": "F", "TTA": "L", "TTG": "L",
    "CTT": "L", "CTC": "L", "CTA": "L", "CTG": "L",
    "ATT": "I", "ATC": "I", "ATA": "I", "ATG": "M",
    "GTT": "V", "GTC": "V", "GTA": "V", "GTG": "V",
    "TCT": "S", "TCC": "S", "TCA": "S", "TCG": "S",
    "CCT": "P", "CCC": "P", "CCA": "P", "CCG": "P",
    "ACT": "T", "ACC": "T", "ACA": "T", "ACG": "T",
    "GCT": "A", "GCC": "A", "GCA": "A", "GCG": "A",
    "TAT": "Y", "TAC": "Y", "TAA": "*", "TAG": "*",
    "CAT": "H", "CAC": "H", "CAA": "Q", "CAG": "Q",
    "AAT": "N", "AAC": "N", "AAA": "K", "AAG": "K",
    "GAT": "D", "GAC": "D", "GAA": "E", "GAG": "E",
    "TGT": "C", "TGC": "C", "TGA": "*", "TGG": "W",
    "CGT": "R", "CGC": "R", "CGA": "R", "CGG": "R",
    "AGT": "S", "AGC": "S", "AGA": "R", "AGG": "R",
    "GGT": "G", "GGC": "G", "GGA": "G", "GGG": "G",
}

_COMPLEMENT_TABLE = str.maketrans("ACGTNacgtn", "TGCANtgcan")


def reverse_complement(seq: str) -> str:
    return seq.translate(_COMPLEMENT_TABLE)[::-1]


def translate_dna(seq: str, *, to_stop: bool = False) -> str:
    """Translate DNA using the standard codon table.

    Unknown codons (non-ACGT or partial trailing codon) are emitted as 'X'.
    `to_stop=True` truncates at the first in-frame stop codon (excluding the
    stop symbol from the output).
    """
    seq = seq.upper()
    out: list[str] = []
    for i in range(0, len(seq) - 2, 3):
        codon = seq[i : i + 3]
        aa = STANDARD_CODON_TABLE.get(codon, "X")
        if aa == "*" and to_stop:
            break
        out.append(aa)
    return "".join(out)


def fasta_wrap(seq: str, width: int = 60) -> str:
    return "\n".join(seq[i : i + width] for i in range(0, len(seq), width))


# ---------------------------------------------------------------------------
# Data preprocessing
# ---------------------------------------------------------------------------

XLSX_NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def parse_xlsx(path: Path) -> list[dict[str, Any]]:
    """Parse the metadata xlsx into a list of records using only stdlib.

    Row 1 is a title row, row 2 is the header. Returns dict per data row.
    """
    with zipfile.ZipFile(path) as z:
        ss_root = ET.fromstring(z.read("xl/sharedStrings.xml"))
        strings = [
            "".join(t.text or "" for t in si.findall(".//a:t", XLSX_NS))
            for si in ss_root.findall("a:si", XLSX_NS)
        ]
        sheet = ET.fromstring(z.read("xl/worksheets/sheet1.xml"))

    def cell_value(c: ET.Element) -> str:
        t = c.attrib.get("t", "")
        v = c.find("a:v", XLSX_NS)
        if t == "inlineStr":
            return "".join(x.text or "" for x in c.findall(".//a:t", XLSX_NS))
        if v is None or v.text is None:
            return ""
        if t == "s":
            return strings[int(v.text)]
        return v.text

    rows = sheet.findall(".//a:row", XLSX_NS)
    if len(rows) < 3:
        return []

    # Row 2 (index 1) is the header per inspection of the file.
    header = [cell_value(c) for c in rows[1].findall("a:c", XLSX_NS)]
    records: list[dict[str, Any]] = []
    for r in rows[2:]:
        cells = [cell_value(c) for c in r.findall("a:c", XLSX_NS)]
        if not cells or not cells[0]:
            continue
        rec = {}
        for i, key in enumerate(header):
            rec[key] = cells[i] if i < len(cells) else ""
        # numeric coercions
        for k in ("Identity", "Kimura.distance"):
            try:
                rec[k] = float(rec.get(k, "")) if rec.get(k, "") != "" else None
            except ValueError:
                pass
        try:
            it = rec.get("Insertion_Time", "")
            rec["Insertion_Time"] = int(float(it)) if it != "" else None
        except ValueError:
            pass
        records.append(rec)
    return records


def _load_genome_info() -> dict[str, dict[str, str]]:
    """Load genome.information.xlsx → {abbr: {full_name, assembly}}.

    The file has a single header row (row 1):
      col 0 = full name (Row name), col 1 = abbreviation, col 2 = assembly/URL
    """
    if not GENOME_INFO_PATH.exists():
        return {}
    try:
        with zipfile.ZipFile(GENOME_INFO_PATH) as z:
            ss_root = ET.fromstring(z.read("xl/sharedStrings.xml"))
            strings = [
                "".join(t.text or "" for t in si.findall(".//a:t", XLSX_NS))
                for si in ss_root.findall("a:si", XLSX_NS)
            ]
            sheet = ET.fromstring(z.read("xl/worksheets/sheet1.xml"))

        def _cv(c: ET.Element) -> str:
            t = c.attrib.get("t", "")
            v = c.find("a:v", XLSX_NS)
            if t == "inlineStr":
                return "".join(x.text or "" for x in c.findall(".//a:t", XLSX_NS))
            if v is None or v.text is None:
                return ""
            if t == "s":
                return strings[int(v.text)]
            return v.text

        rows = sheet.findall(".//a:row", XLSX_NS)
        result: dict[str, dict[str, str]] = {}
        for row in rows[1:]:  # skip header row
            cells = [_cv(c) for c in row.findall("a:c", XLSX_NS)]
            if len(cells) < 2:
                continue
            full_name = cells[0].strip() if cells[0] else ""
            abbr = cells[1].strip() if cells[1] else ""
            assembly = cells[2].strip() if len(cells) > 2 and cells[2] else ""
            # strip trailing tab/whitespace from assembly
            assembly = assembly.rstrip("\t")
            if abbr:
                result[abbr] = {"full_name": full_name or abbr, "assembly": assembly}
        return result
    except Exception as exc:  # noqa: BLE001
        print(f"[WARN] Could not load genome.information.xlsx: {exc}")
        return {}


GENOME_INFO: dict[str, dict[str, str]] = _load_genome_info()


def parse_bed(path: Path) -> dict[str, list[dict[str, Any]]]:
    out: dict[str, list[dict[str, Any]]] = defaultdict(list)
    with path.open("r", encoding="utf-8") as fh:
        for raw in fh:
            line = raw.rstrip("\n")
            if not line or line.startswith("#"):
                continue
            parts = line.split("\t")
            if len(parts) < 4:
                continue
            sid = parts[0]
            try:
                start = int(parts[1])
                end = int(parts[2])
            except ValueError:
                continue
            name = parts[3]
            strand = parts[5] if len(parts) > 5 and parts[5] in {"+", "-"} else "+"
            out[sid].append(
                {"name": name, "start": start, "end": end, "strand": strand}
            )
    for sid in out:
        out[sid].sort(key=lambda r: (r["start"], r["end"]))
    return dict(out)


def build_fasta_offsets(path: Path) -> dict[str, dict[str, int]]:
    """Index a FASTA file into {seq_id: {seq_start, seq_end, length}} byte offsets.

    `seq_start` points to the first byte of sequence data after the header line;
    `seq_end` points to the byte right after the last sequence byte (exclusive).
    Sequence lines may contain newlines that we strip when reading.
    """
    index: dict[str, dict[str, int]] = {}
    with path.open("rb") as fh:
        data = fh.read()
    pos = 0
    n = len(data)
    while pos < n:
        if data[pos : pos + 1] != b">":
            # find next header
            nxt = data.find(b"\n>", pos)
            if nxt == -1:
                break
            pos = nxt + 1
            continue
        header_end = data.find(b"\n", pos)
        if header_end == -1:
            break
        header = data[pos + 1 : header_end].decode("utf-8", errors="replace").strip()
        sid = header.split()[0] if header else ""
        seq_start = header_end + 1
        next_header = data.find(b"\n>", seq_start)
        if next_header == -1:
            seq_end = n
            pos = n
        else:
            seq_end = next_header  # newline before '>'
            pos = next_header + 1
        # compute pure sequence length without newlines / whitespace
        block = data[seq_start:seq_end]
        seq_len = sum(1 for b in block if b not in (10, 13, 32, 9))
        if sid:
            index[sid] = {
                "seq_start": seq_start,
                "seq_end": seq_end,
                "length": seq_len,
            }
    return index


def read_fasta_full(path: Path, offsets: dict[str, int]) -> str:
    """Read the full sequence (concatenated, whitespace stripped) from byte offsets."""
    with path.open("rb") as fh:
        fh.seek(offsets["seq_start"])
        block = fh.read(offsets["seq_end"] - offsets["seq_start"])
    return re.sub(r"\s+", "", block.decode("ascii", errors="replace")).upper()


# ---------------------------------------------------------------------------
# Build / load index files
# ---------------------------------------------------------------------------

def _needs_rebuild(target: Path, sources: list[Path]) -> bool:
    if not target.exists():
        return True
    target_mtime = target.stat().st_mtime
    return any(s.exists() and s.stat().st_mtime > target_mtime for s in sources)


def build_indexes(force: bool = False) -> None:
    if force or _needs_rebuild(META_JSON, [XLSX_PATH]):
        meta = parse_xlsx(XLSX_PATH)
        META_JSON.write_text(json.dumps(meta, ensure_ascii=False))
    if force or _needs_rebuild(ORF_JSON, [ORF_BED_PATH]):
        ORF_JSON.write_text(json.dumps(parse_bed(ORF_BED_PATH), ensure_ascii=False))
    if force or _needs_rebuild(DOMAIN_JSON, [DOMAIN_BED_PATH]):
        DOMAIN_JSON.write_text(
            json.dumps(parse_bed(DOMAIN_BED_PATH), ensure_ascii=False)
        )
    if force or _needs_rebuild(SEQ_OFFSETS_JSON, [PASS_FA_PATH]):
        SEQ_OFFSETS_JSON.write_text(
            json.dumps(build_fasta_offsets(PASS_FA_PATH), ensure_ascii=False)
        )


def _load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


# ---------------------------------------------------------------------------
# Flask app
# ---------------------------------------------------------------------------

app = Flask(__name__, static_folder="static", template_folder="templates")

# build indexes once on import (safe across worker processes; idempotent)
build_indexes(force=False)

META: list[dict[str, Any]] = _load_json(META_JSON)
ORF_INDEX: dict[str, list[dict[str, Any]]] = _load_json(ORF_JSON)
DOMAIN_INDEX: dict[str, list[dict[str, Any]]] = _load_json(DOMAIN_JSON)
SEQ_OFFSETS: dict[str, dict[str, int]] = _load_json(SEQ_OFFSETS_JSON)


# ---------------------------------------------------------------------------
# Genome-browser indexes (lazily loaded; built by build_genome_index.py)
# ---------------------------------------------------------------------------

_GENOME_FAI: dict[str, dict] | None = None


def _genome_ready() -> tuple[bool, str]:
    if not GENOME_FASTA.exists():
        return False, f"reference FASTA missing: {GENOME_FASTA}"
    if not GENOME_FAI.exists():
        return False, "FASTA index missing; run: python build_genome_index.py"
    if not GENOME_DB.exists():
        return False, "GTF SQLite missing; run: python build_genome_index.py"
    # Whole-genome BED12 is built lazily on first /genome request - we only
    # signal readiness based on the heavy artefacts above.
    return True, ""


# Stable per-biotype palette so different transcript classes get distinct
# colours in the IGV BED12 track. Falls back to the default protein_coding
# blue for anything unknown.
_BIOTYPE_RGB = {
    "protein_coding": "37,99,235",        # blue
    "lncRNA": "168,85,247",               # purple
    "miRNA": "236,72,153",                # pink
    "snoRNA": "236,72,153",
    "snRNA": "236,72,153",
    "rRNA": "234,88,12",                  # orange
    "tRNA": "234,88,12",
    "Mt_rRNA": "234,88,12",
    "Mt_tRNA": "234,88,12",
    "processed_pseudogene": "148,163,184",  # slate
    "pseudogene": "148,163,184",
    "transcribed_processed_pseudogene": "148,163,184",
    "transcribed_unprocessed_pseudogene": "148,163,184",
    "unprocessed_pseudogene": "148,163,184",
    "retained_intron": "20,184,166",       # teal
    "processed_transcript": "20,184,166",
    "nonsense_mediated_decay": "239,68,68",
}


def _build_genome_bed() -> None:
    """Generate transcript-level BED12 plus gene-level BED9 from the GTF.

    Two static files drive the IGV browser:

    * `data/genome.bed`        - one BED12 record per transcript with the
      complete exon/CDS structure. igv.js's native BED parser stacks each
      isoform onto its own row in EXPANDED mode and renders thick CDS /
      thin UTR boxes for free.
    * `data/genome.genes.bed`  - one BED9 record per gene (plain box
      spanning the gene locus). Gives users a quick "where does this
      gene live" overview that doesn't change with isoform layout.

    The transcript BED uses `transcript_id` as the BED `name` column. We
    deliberately do NOT embed the gene symbol there - some BED tokenisers
    split on whitespace, so `"ENSSSCT00000027607 (ALDH1A1)"` was being
    parsed as a BED6 row and the exon block columns went missing, which is
    exactly why earlier builds rendered each transcript as a flat box with
    no exon/intron structure. Gene symbols stay reachable via the right-
    side detail panel.
    """
    conn = _get_gtf_conn()
    try:
        rows = conn.execute(
            "SELECT type, chrom, start, end, strand, transcript_id, gene_name, "
            "gene_id, transcript_biotype, gene_biotype "
            "FROM features WHERE type IN ('transcript','exon','CDS') "
            "ORDER BY chrom, start, end"
        ).fetchall()
        gene_rows = conn.execute(
            "SELECT chrom, start, end, strand, gene_id, gene_name, gene_biotype "
            "FROM features WHERE type='gene' ORDER BY chrom, start, end"
        ).fetchall()
    finally:
        conn.close()

    txs: dict[str, dict[str, Any]] = {}
    for r in rows:
        tid = r["transcript_id"]
        if not tid:
            continue
        t = txs.get(tid)
        if t is None:
            t = {
                "chrom": str(r["chrom"] or ""),
                "name": str(tid),
                "strand": str(r["strand"] or "."),
                "biotype": str(r["transcript_biotype"] or r["gene_biotype"] or ""),
                "tx_start": None,
                "tx_end": None,
                "exons": [],
                "cds": [],
            }
            txs[tid] = t
        rs = int(r["start"]) - 1
        re_ = int(r["end"])
        rt = r["type"]
        if rt == "transcript":
            t["tx_start"] = rs
            t["tx_end"] = re_
        elif rt == "exon":
            t["exons"].append((rs, re_))
        elif rt == "CDS":
            t["cds"].append((rs, re_))

    GENOME_BED.parent.mkdir(parents=True, exist_ok=True)
    tmp = GENOME_BED.with_suffix(".bed.tmp")
    with tmp.open("w", encoding="utf-8") as fh:
        for t in txs.values():
            if not t["exons"] and t["tx_start"] is None:
                continue
            if t["tx_start"] is None or t["tx_end"] is None:
                t["tx_start"] = min(e[0] for e in t["exons"]) if t["exons"] else 0
                t["tx_end"] = max(e[1] for e in t["exons"]) if t["exons"] else 0
            cstart = t["tx_start"]
            cend = t["tx_end"]
            exons = sorted(t["exons"]) or [(cstart, cend)]
            if t["cds"]:
                thick_start = min(c[0] for c in t["cds"])
                thick_end = max(c[1] for c in t["cds"])
            else:
                thick_start = cend
                thick_end = cend
            block_count = len(exons)
            block_sizes = ",".join(str(e[1] - e[0]) for e in exons) + ","
            block_starts = ",".join(str(e[0] - cstart) for e in exons) + ","
            # itemRgb is intentionally "0" (no per-record colour) so that the
            # IGV track-level colour picker / `track.color` config controls
            # the whole track uniformly. Earlier builds wrote a biotype-keyed
            # RGB here, which forced every record into a fixed colour and
            # made the IGV "Set track colour" picker only seem to affect
            # records of unknown biotype.
            fh.write(
                "\t".join(
                    [
                        t["chrom"],
                        str(cstart),
                        str(cend),
                        t["name"],  # transcript_id only - no spaces, parses as BED12
                        "0",
                        t["strand"],
                        str(thick_start),
                        str(thick_end),
                        "0",
                        str(block_count),
                        block_sizes,
                        block_starts,
                    ]
                )
                + "\n"
            )
    tmp.replace(GENOME_BED)

    # gene-level BED (BED9: chrom, start, end, name, score, strand, thickStart,
    # thickEnd, itemRgb). Genes don't need block info - they're a flat span.
    # itemRgb is "0" on purpose so the IGV track colour picker / track.color
    # config controls the entire track uniformly.
    gene_tmp = GENOME_GENES_BED.with_suffix(".bed.tmp")
    with gene_tmp.open("w", encoding="utf-8") as fh:
        for r in gene_rows:
            cstart = max(0, int(r["start"]) - 1)
            cend = int(r["end"])
            label = str(r["gene_name"] or r["gene_id"] or "gene")
            label = label.replace(" ", "_")  # be safe even though Ensembl ids are clean
            fh.write(
                "\t".join(
                    [
                        str(r["chrom"]),
                        str(cstart),
                        str(cend),
                        label,
                        "0",
                        str(r["strand"] or "."),
                        str(cstart),
                        str(cend),
                        "0",
                    ]
                )
                + "\n"
            )
    gene_tmp.replace(GENOME_GENES_BED)


def _ensure_genome_bed() -> None:
    """(Re)build BED files if they are missing or older than the SQLite index."""
    if not GENOME_DB.exists():
        return
    db_mtime = GENOME_DB.stat().st_mtime
    needs = (
        not GENOME_BED.exists()
        or GENOME_BED.stat().st_mtime < db_mtime
        or not GENOME_GENES_BED.exists()
        or GENOME_GENES_BED.stat().st_mtime < db_mtime
    )
    if needs:
        _build_genome_bed()


def _ensure_perv_bed() -> None:
    """Generate data/perv.bed from Homologous/RF.intact.region if missing or stale.

    RF.intact.region is 1-based closed interval (non-BED). To convert to BED6:
      bed_start = col2 - 1,  bed_end = col3  (unchanged)
    """
    if not PERV_REGION_FILE.exists():
        return
    src_mtime = PERV_REGION_FILE.stat().st_mtime
    if PERV_BED.exists() and PERV_BED.stat().st_mtime >= src_mtime:
        return
    lines: list[str] = []
    with PERV_REGION_FILE.open() as fh:
        for raw in fh:
            raw = raw.rstrip("\n")
            if not raw or raw.startswith("#"):
                continue
            parts = raw.split("\t")
            if len(parts) < 6:
                continue
            chrom, s, e, name, score, strand = parts[0], parts[1], parts[2], parts[3], parts[4], parts[5]
            try:
                bed_start = int(s) - 1  # 1-based → 0-based
                bed_end = int(e)        # 1-based closed == 0-based exclusive
            except ValueError:
                continue
            lines.append(f"{chrom}\t{bed_start}\t{bed_end}\t{name}\t0\t{strand}")
    PERV_BED.write_text("\n".join(lines) + "\n")


# Cache for PERV list JSON (loaded once per process).
_PERV_LIST_CACHE: list | None = None


def _load_perv_list() -> list[dict]:
    """Parse RF.intact.region + domain/ORF BED files, return transformed list.

    RF.intact.region: 1-based closed interval [start, end] (NOT BED).
    BED files: 0-based half-open [start, end).

    Coordinate transform to genomic BED coords:
      + strand: abs_start = (region_start - 1) + rel_start
                abs_end   = (region_start - 1) + rel_end
      - strand: abs_start = region_end - rel_end
                abs_end   = region_end - rel_start
    """
    global _PERV_LIST_CACHE
    if _PERV_LIST_CACHE is not None:
        return _PERV_LIST_CACHE

    # --- parse RF.intact.region ---
    seqs: dict[str, dict] = {}
    if PERV_REGION_FILE.exists():
        with PERV_REGION_FILE.open() as fh:
            for raw in fh:
                raw = raw.rstrip("\n")
                if not raw or raw.startswith("#"):
                    continue
                parts = raw.split("\t")
                if len(parts) < 6:
                    continue
                chrom, s, e, name, _score, strand = parts[0], parts[1], parts[2], parts[3], parts[4], parts[5]
                try:
                    start1 = int(s)  # 1-based
                    end1 = int(e)    # 1-based closed
                except ValueError:
                    continue
                seqs[name] = {
                    "name": name,
                    "chrom": chrom,
                    "start": start1,    # keep 1-based for display
                    "end": end1,        # 1-based closed
                    "strand": strand,
                    "length": end1 - start1 + 1,
                    "domains": [],
                    "orfs": [],
                }

    def _transform(region: dict, rel_start: int, rel_end: int) -> tuple[int, int]:
        """Convert relative BED coords to absolute genomic BED coords (0-based half-open)."""
        s1 = region["start"]   # 1-based
        e1 = region["end"]     # 1-based closed (= 0-based exclusive)
        if region["strand"] == "+":
            return (s1 - 1) + rel_start, (s1 - 1) + rel_end
        else:
            return e1 - rel_end, e1 - rel_start

    # --- parse domain BED ---
    if DOMAIN_BED_PATH.exists():
        with DOMAIN_BED_PATH.open() as fh:
            for raw in fh:
                raw = raw.rstrip("\n")
                if not raw or raw.startswith("#"):
                    continue
                parts = raw.split("\t")
                if len(parts) < 4:
                    continue
                seq_id = parts[0]
                if seq_id not in seqs:
                    continue
                try:
                    rel_s, rel_e = int(parts[1]), int(parts[2])
                except ValueError:
                    continue
                feat_name = parts[3]
                abs_s, abs_e = _transform(seqs[seq_id], rel_s, rel_e)
                seqs[seq_id]["domains"].append({
                    "name": feat_name,
                    "start": abs_s,
                    "end": abs_e,
                    "strand": seqs[seq_id]["strand"],
                    "length": abs_e - abs_s,
                })

    # --- parse ORF BED ---
    if ORF_BED_PATH.exists():
        with ORF_BED_PATH.open() as fh:
            for raw in fh:
                raw = raw.rstrip("\n")
                if not raw or raw.startswith("#"):
                    continue
                parts = raw.split("\t")
                if len(parts) < 4:
                    continue
                seq_id = parts[0]
                if seq_id not in seqs:
                    continue
                try:
                    rel_s, rel_e = int(parts[1]), int(parts[2])
                except ValueError:
                    continue
                feat_name = parts[3]
                abs_s, abs_e = _transform(seqs[seq_id], rel_s, rel_e)
                seqs[seq_id]["orfs"].append({
                    "name": feat_name,
                    "start": abs_s,
                    "end": abs_e,
                    "strand": seqs[seq_id]["strand"],
                    "length": abs_e - abs_s,
                })

    # --- annotate ERV.type from 1165.intact.PERV.infomation.xlsx ---
    if XLSX_PATH.exists():
        try:
            import openpyxl as _openpyxl
            _wb = _openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
            _ws = _wb.active
            _rows = _ws.iter_rows(values_only=True)
            next(_rows)  # skip title row
            _hdrs = [str(h).strip() if h is not None else "" for h in next(_rows)]
            _id_col = next((i for i, h in enumerate(_hdrs) if h == "Sequence.ID"), None)
            _type_col = next((i for i, h in enumerate(_hdrs) if h in ("ERV.type", "ERV_type")), None)
            if _id_col is not None and _type_col is not None:
                for _row in _rows:
                    _sid = str(_row[_id_col]).strip() if _row[_id_col] is not None else ""
                    _etype = str(_row[_type_col]).strip() if _row[_type_col] is not None else ""
                    if _sid in seqs and _etype:
                        seqs[_sid]["erv_type"] = _etype
            _wb.close()
        except Exception:
            pass  # non-critical; gracefully skip if xlsx unavailable

    result = list(seqs.values())
    _PERV_LIST_CACHE = result
    return result


# ---------------------------------------------------------------------------
# Homologous sequences (final.Statistics.table.xlsx)
# ---------------------------------------------------------------------------

_HOMOLOGOUS_CACHE: tuple | None = None
_HOMOLOGOUS_CACHE_MTIME: float | None = None


def _parse_homologous_xlsx() -> list[dict]:
    """Parse final.Statistics.table.xlsx (row 1 = header, numeric coords in cols 2,3,10,11)."""
    with zipfile.ZipFile(HOMOLOGOUS_XLSX) as z:
        ss_root = ET.fromstring(z.read("xl/sharedStrings.xml"))
        strings = [
            "".join(t.text or "" for t in si.findall(".//a:t", XLSX_NS))
            for si in ss_root.findall("a:si", XLSX_NS)
        ]
        sheet = ET.fromstring(z.read("xl/worksheets/sheet1.xml"))

    def cell_val(c: ET.Element) -> str:
        t = c.attrib.get("t", "")
        v = c.find("a:v", XLSX_NS)
        if t == "inlineStr":
            return "".join(x.text or "" for x in c.findall(".//a:t", XLSX_NS))
        if v is None or v.text is None:
            return ""
        if t == "s":
            return strings[int(v.text)]
        return v.text

    rows = sheet.findall(".//a:row", XLSX_NS)
    if len(rows) < 2:
        return []

    header = [cell_val(c) for c in rows[0].findall("a:c", XLSX_NS)]
    records: list[dict] = []
    for r in rows[1:]:
        cells = [cell_val(c) for c in r.findall("a:c", XLSX_NS)]
        if not cells or not cells[0]:
            continue
        rec: dict = {}
        for i, key in enumerate(header):
            rec[key] = cells[i] if i < len(cells) else ""
        for k in ("new.start", "new.end", "locus_start", "locus_end"):
            try:
                val = rec.get(k, "")
                rec[k] = int(float(val)) if val != "" else None
            except (ValueError, TypeError):
                rec[k] = None
        records.append(rec)
    return records


def _load_homologous() -> tuple[list[dict], dict[str, dict]]:
    """Load and cache homologous sequence data from Excel.

    Returns:
        (seqs, loci_map) where
        - seqs      : list of 876 dicts, Chr field prefixed with 'chr', coords 1-based
        - loci_map  : dict locus_id → locus info + species/group aggregation
    """
    global _HOMOLOGOUS_CACHE, _HOMOLOGOUS_CACHE_MTIME

    if not HOMOLOGOUS_XLSX.exists():
        _HOMOLOGOUS_CACHE = ([], {})
        _HOMOLOGOUS_CACHE_MTIME = None
        return _HOMOLOGOUS_CACHE

    src_mtime = HOMOLOGOUS_XLSX.stat().st_mtime
    if _HOMOLOGOUS_CACHE is not None and _HOMOLOGOUS_CACHE_MTIME == src_mtime:
        return _HOMOLOGOUS_CACHE

    # Build ERV type lookup from PERV xlsx (q_name → erv_type)
    _erv_type_map: dict[str, str] = {}
    if XLSX_PATH.exists():
        try:
            import openpyxl as _openpyxl
            _wb = _openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
            _ws = _wb.active
            _irows = _ws.iter_rows(values_only=True)
            next(_irows)  # skip title row
            _hdrs = [str(h).strip() if h is not None else "" for h in next(_irows)]
            _id_col = next((i for i, h in enumerate(_hdrs) if h == "Sequence.ID"), None)
            _type_col = next((i for i, h in enumerate(_hdrs) if h in ("ERV.type", "ERV_type")), None)
            if _id_col is not None and _type_col is not None:
                for _row in _irows:
                    _sid = str(_row[_id_col]).strip() if _row[_id_col] is not None else ""
                    _et = str(_row[_type_col]).strip() if _row[_type_col] is not None else ""
                    if _sid and _et:
                        _erv_type_map[_sid] = _et
            _wb.close()
        except Exception:
            pass

    raw = _parse_homologous_xlsx()
    seqs: list[dict] = []
    loci_map: dict[str, dict] = {}

    for r in raw:
        chrom = "chr" + str(r.get("Chr", "")) if r.get("Chr") else ""
        qname = r.get("q_name", "")
        seq = {
            "q_name":       qname,
            "start":        r.get("new.start"),   # 1-based, keep for display
            "end":          r.get("new.end"),      # 1-based closed
            "strand":       r.get("strand_up", "+"),
            "species":      r.get("species", ""),
            "chrom":        chrom,
            "group":        r.get("group", ""),
            "locus_id":     r.get("locus_id", ""),
            "locus_label":  r.get("locus_label", ""),
            "locus_start":  r.get("locus_start"),
            "locus_end":    r.get("locus_end"),
            "erv_type":     _erv_type_map.get(qname, ""),
        }
        seqs.append(seq)

        lid = seq["locus_id"]
        if lid not in loci_map:
            locus_label = r.get("locus_label", "")
            m = re.match(r'^([^:]+):(\d+)-(\d+)\(([+-])\)', locus_label)
            locus_chrom = ("chr" + m.group(1)) if m else chrom
            locus_strand = m.group(4) if m else "+"
            loci_map[lid] = {
                "locus_id":    lid,
                "chrom":       locus_chrom,
                "start":       seq["locus_start"],  # 1-based
                "end":         seq["locus_end"],     # 1-based closed
                "strand":      locus_strand,
                "count":       0,
                "species_dist": {},
                "group_dist":  {},
            }
        loci_map[lid]["count"] += 1
        sp = seq["species"]
        grp = seq["group"]
        loci_map[lid]["species_dist"][sp] = loci_map[lid]["species_dist"].get(sp, 0) + 1
        loci_map[lid]["group_dist"][grp] = loci_map[lid]["group_dist"].get(grp, 0) + 1

    _HOMOLOGOUS_CACHE = (seqs, loci_map)
    _HOMOLOGOUS_CACHE_MTIME = src_mtime
    return _HOMOLOGOUS_CACHE


def _ensure_homologous_beds() -> None:
    """Generate BED6 files for homologous sequences and loci if missing/stale."""
    if not HOMOLOGOUS_XLSX.exists():
        return
    src_mtime = HOMOLOGOUS_XLSX.stat().st_mtime
    both_fresh = (
        HOMOLOGOUS_SEQ_BED.exists() and HOMOLOGOUS_SEQ_BED.stat().st_mtime >= src_mtime
        and HOMOLOGOUS_LOCUS_BED.exists() and HOMOLOGOUS_LOCUS_BED.stat().st_mtime >= src_mtime
    )
    if both_fresh:
        return

    seqs, loci_map = _load_homologous()

    seq_lines: list[str] = []
    for s in seqs:
        if s["start"] is None or s["end"] is None:
            continue
        bed_start = s["start"] - 1   # 1-based → 0-based
        bed_end = s["end"]            # 1-based closed == 0-based exclusive
        strand = s["strand"] or "+"
        seq_lines.append(f"{s['chrom']}\t{bed_start}\t{bed_end}\t{s['q_name']}\t0\t{strand}")
    HOMOLOGOUS_SEQ_BED.write_text("\n".join(seq_lines) + "\n")

    locus_lines: list[str] = []
    for lid, locus in sorted(loci_map.items()):
        if locus["start"] is None or locus["end"] is None:
            continue
        bed_start = locus["start"] - 1
        bed_end = locus["end"]
        strand = locus["strand"] or "+"
        locus_lines.append(
            f"{locus['chrom']}\t{bed_start}\t{bed_end}\t{lid}\t0\t{strand}"
        )
    HOMOLOGOUS_LOCUS_BED.write_text("\n".join(locus_lines) + "\n")


def _get_fai() -> dict[str, dict]:
    global _GENOME_FAI
    if _GENOME_FAI is None:
        _GENOME_FAI = genome_lib.load_fai(GENOME_FAI)
    return _GENOME_FAI


def _get_gtf_conn():
    return genome_lib.gtf_connect(GENOME_DB)


def _require_genome():
    ok, msg = _genome_ready()
    if not ok:
        abort(503, description=msg)


# Static-asset cache busting: use the process start time so each restart
# forces browsers to refetch JS/CSS instead of running stale cached copies.
_ASSET_VERSION = str(int(time.time()))


@app.context_processor
def _inject_asset_version():
    return {"asset_v": _ASSET_VERSION}


# -- pages ------------------------------------------------------------------

@app.route("/")
def page_index():
    return render_template("index.html")


@app.route("/overview")
def page_overview():
    return render_template("overview.html")


@app.route("/browser")
def page_browser():
    return render_template("browser.html")


@app.route("/genome")
def page_genome():
    ok, msg = _genome_ready()
    return render_template("genome.html", genome_ready=ok, genome_msg=msg)


# -- overview APIs ----------------------------------------------------------

def _hist(values: list[float], bins: int = 20) -> dict[str, list]:
    vals = [v for v in values if isinstance(v, (int, float))]
    if not vals:
        return {"edges": [], "counts": []}
    lo, hi = min(vals), max(vals)
    if lo == hi:
        return {"edges": [lo, hi + 1e-9], "counts": [len(vals)]}
    width = (hi - lo) / bins
    edges = [lo + i * width for i in range(bins + 1)]
    counts = [0] * bins
    for v in vals:
        idx = int((v - lo) / width)
        if idx == bins:
            idx = bins - 1
        counts[idx] += 1
    return {"edges": edges, "counts": counts}


@app.route("/api/overview/stats")
def api_overview_stats():
    total = len(META)
    type_counter = Counter(r.get("ERV.type", "") for r in META)
    group_counter = Counter(r.get("Group", "") for r in META)
    abbr_counter = Counter(r.get("Abbretiation", "") for r in META)

    identity_vals = [r.get("Identity") for r in META if r.get("Identity") is not None]
    insertion_vals = [
        r.get("Insertion_Time") for r in META if r.get("Insertion_Time") is not None
    ]
    kimura_vals = [
        r.get("Kimura.distance") for r in META if r.get("Kimura.distance") is not None
    ]

    abbr_sorted = sorted(abbr_counter.items(), key=lambda kv: -kv[1])
    return jsonify(
        {
            "total": total,
            "type_counts": dict(type_counter),
            "group_counts": dict(group_counter),
            "abbr_counts": [
                {
                    "name": k,
                    "count": v,
                    "full_name": GENOME_INFO.get(k, {}).get("full_name", k),
                    "assembly": GENOME_INFO.get(k, {}).get("assembly", ""),
                }
                for k, v in abbr_sorted
            ],
            "identity_hist": _hist(identity_vals, bins=20),
            "insertion_hist": _hist(insertion_vals, bins=20),
            "kimura_hist": _hist(kimura_vals, bins=20),
        }
    )


@app.route("/api/genome/genome_info")
def api_genome_info():
    """Return genome info mapping: abbr → {full_name, assembly}."""
    return jsonify(GENOME_INFO)


@app.route("/api/overview/table")
def api_overview_table():
    q = request.args.get("q", "").strip().lower()
    erv_type = request.args.get("type", "").strip()
    group = request.args.get("group", "").strip()
    page = max(1, int(request.args.get("page", "1") or "1"))
    size = max(1, min(200, int(request.args.get("size", "25") or "25")))

    rows = META
    if q:
        rows = [r for r in rows if q in str(r.get("Sequence.ID", "")).lower()]
    if erv_type:
        rows = [r for r in rows if r.get("ERV.type") == erv_type]
    if group:
        rows = [r for r in rows if r.get("Group") == group]

    total = len(rows)
    start = (page - 1) * size
    end = start + size
    return jsonify(
        {
            "total": total,
            "page": page,
            "size": size,
            "rows": rows[start:end],
        }
    )


# -- pass.139 sequence APIs --------------------------------------------------

@app.route("/api/sequences/pass")
def api_sequences_pass():
    items = [
        {"id": sid, "length": SEQ_OFFSETS[sid]["length"]}
        for sid in sorted(SEQ_OFFSETS.keys())
    ]
    return jsonify({"total": len(items), "items": items})


def _get_full_seq(sid: str) -> str:
    if sid not in SEQ_OFFSETS:
        abort(404, description=f"sequence {sid} not found")
    return read_fasta_full(PASS_FA_PATH, SEQ_OFFSETS[sid])


@app.route("/api/sequences/<sid>/regions")
def api_sequence_regions(sid: str):
    kind = request.args.get("kind", "orf").lower()
    if kind == "orf":
        regions = ORF_INDEX.get(sid, [])
    elif kind == "domain":
        regions = DOMAIN_INDEX.get(sid, [])
    else:
        abort(400, description="kind must be 'orf' or 'domain'")
    if sid not in SEQ_OFFSETS:
        abort(404, description=f"sequence {sid} not found")
    return jsonify(
        {
            "id": sid,
            "length": SEQ_OFFSETS[sid]["length"],
            "kind": kind,
            "regions": regions,
        }
    )


def _slice_region(sid: str, start: int, end: int, strand: str) -> str:
    full = _get_full_seq(sid)
    if start < 0 or end > len(full) or start >= end:
        abort(400, description="invalid start/end")
    sub = full[start:end]
    if strand == "-":
        sub = reverse_complement(sub)
    return sub


@app.route("/api/sequences/<sid>/dna")
def api_sequence_dna(sid: str):
    try:
        start = int(request.args["start"])
        end = int(request.args["end"])
    except (KeyError, ValueError):
        abort(400, description="start and end query params required")
    strand = request.args.get("strand", "+")
    name = request.args.get("name", "")
    sub = _slice_region(sid, start, end, strand)
    full_len = SEQ_OFFSETS[sid]["length"]
    return jsonify(
        {
            "id": sid,
            "name": name,
            "start": start,
            "end": end,
            "strand": strand,
            "length": end - start,
            "full_length": full_len,
            "dna": sub,
            "fasta": f">{sid}|{name}|{start}-{end}|{strand}\n{fasta_wrap(sub)}\n",
        }
    )


@app.route("/api/sequences/<sid>/protein")
def api_sequence_protein(sid: str):
    try:
        start = int(request.args["start"])
        end = int(request.args["end"])
    except (KeyError, ValueError):
        abort(400, description="start and end query params required")
    strand = request.args.get("strand", "+")
    name = request.args.get("name", "")
    sub = _slice_region(sid, start, end, strand)
    protein = translate_dna(sub, to_stop=False)
    return jsonify(
        {
            "id": sid,
            "name": name,
            "start": start,
            "end": end,
            "strand": strand,
            "length": len(protein),
            "protein": protein,
            "fasta": f">{sid}|{name}|{start}-{end}|{strand}|protein\n{fasta_wrap(protein)}\n",
        }
    )


@app.route("/api/sequences/<sid>/all-protein")
def api_sequence_all_protein(sid: str):
    kind = request.args.get("kind", "orf").lower()
    if kind == "orf":
        regions = [r for r in ORF_INDEX.get(sid, []) if r["name"] != "LTR"]
    elif kind == "domain":
        regions = list(DOMAIN_INDEX.get(sid, []))
    else:
        abort(400, description="kind must be 'orf' or 'domain'")
    full = _get_full_seq(sid)
    chunks: list[str] = []
    items = []
    for r in regions:
        sub = full[r["start"] : r["end"]]
        if r.get("strand", "+") == "-":
            sub = reverse_complement(sub)
        protein = translate_dna(sub)
        chunks.append(
            f">{sid}|{r['name']}|{r['start']}-{r['end']}|{r.get('strand', '+')}|protein\n"
            + fasta_wrap(protein)
            + "\n"
        )
        items.append(
            {
                "name": r["name"],
                "start": r["start"],
                "end": r["end"],
                "strand": r.get("strand", "+"),
                "protein": protein,
            }
        )
    return jsonify({"id": sid, "kind": kind, "items": items, "fasta": "".join(chunks)})


# -- genome browser APIs ----------------------------------------------------

# Sequences with these names appear first in chromosome lists / dropdowns.
_PRIMARY_CHROMS = [f"chr{i}" for i in range(1, 19)] + ["chrX", "chrY", "chrM"]


@app.route("/api/genome/status")
def api_genome_status():
    ok, msg = _genome_ready()
    return jsonify({"ready": ok, "message": msg})


@app.route("/api/genome/chromosomes")
def api_genome_chromosomes():
    _require_genome()
    fai = _get_fai()
    items = [{"name": k, "length": v["length"]} for k, v in fai.items()]
    primary_set = set(_PRIMARY_CHROMS)
    primary_order = {n: i for i, n in enumerate(_PRIMARY_CHROMS)}

    def sort_key(it):
        n = it["name"]
        if n in primary_set:
            return (0, primary_order[n])
        return (1, n)

    items.sort(key=sort_key)
    return jsonify({"total": len(items), "items": items})


def _parse_region_args() -> tuple[str, int, int]:
    chrom = request.args.get("chrom", "").strip()
    if not chrom:
        abort(400, description="chrom is required")
    fai = _get_fai()
    if chrom not in fai:
        abort(404, description=f"chromosome {chrom} not found")
    chrom_len = fai[chrom]["length"]
    try:
        start = int(request.args.get("start", "1"))
        end = int(request.args.get("end", str(chrom_len)))
    except ValueError:
        abort(400, description="start/end must be integers")
    if start < 1:
        start = 1
    if end > chrom_len:
        end = chrom_len
    if end < start:
        abort(400, description="end must be >= start")
    return chrom, start, end


@app.route("/api/genome/sequence")
def api_genome_sequence():
    _require_genome()
    chrom, start, end = _parse_region_args()
    span = end - start + 1
    if span > MAX_DNA_REGION:
        abort(
            400,
            description=f"region too large ({span} bp); max is {MAX_DNA_REGION} bp. "
            "Zoom in or use the export endpoint.",
        )
    seq = genome_lib.read_fasta_region(GENOME_FASTA, _get_fai(), chrom, start, end)
    return jsonify(
        {
            "chrom": chrom,
            "start": start,
            "end": end,
            "length": len(seq),
            "sequence": seq,
        }
    )


def _row_to_feature(r) -> dict[str, Any]:
    return {
        "chrom": r["chrom"],
        "type": r["type"],
        "start": r["start"],
        "end": r["end"],
        "strand": r["strand"],
        "phase": r["phase"],
        "gene_id": r["gene_id"],
        "transcript_id": r["transcript_id"],
        "gene_name": r["gene_name"],
        "gene_biotype": r["gene_biotype"],
        "transcript_biotype": r["transcript_biotype"],
    }


@app.route("/api/genome/features")
def api_genome_features():
    _require_genome()
    chrom, start, end = _parse_region_args()
    types_arg = request.args.get("types", "gene,transcript,exon,CDS")
    types = [t.strip() for t in types_arg.split(",") if t.strip()]
    try:
        limit = min(MAX_FEATURES, max(1, int(request.args.get("limit", str(MAX_FEATURES)))))
    except ValueError:
        limit = MAX_FEATURES
    conn = _get_gtf_conn()
    try:
        rows = genome_lib.features_in_region(conn, chrom, start, end, types, limit)
    finally:
        conn.close()
    return jsonify(
        {
            "chrom": chrom,
            "start": start,
            "end": end,
            "types": types,
            "limit": limit,
            "count": len(rows),
            "truncated": len(rows) >= limit,
            "features": [_row_to_feature(r) for r in rows],
        }
    )


@app.route("/api/genome/igv")
def api_genome_igv():
    """IGV-friendly features for the annotation track.

    - At large span (> 5 Mb): one record per gene (light, fast overview).
    - Otherwise: one record per transcript with grouped `exons` (with cdStart/cdEnd
      derived from CDS rows) so igv.js draws proper gene structures with introns
      and thicker CDS portions.
    """
    _require_genome()
    chrom, start, end = _parse_region_args()
    span = end - start + 1

    conn = _get_gtf_conn()
    try:
        if span > 5_000_000:
            rows = genome_lib.features_in_region(
                conn, chrom, start, end, ["gene"], MAX_FEATURES
            )
            features = [
                {
                    "chr": str(r["chrom"] or ""),
                    "start": int(r["start"]) - 1,
                    "end": int(r["end"]),
                    "name": str(r["gene_name"] or r["gene_id"] or ""),
                    "strand": str(r["strand"] or "."),
                    "gene_id": str(r["gene_id"] or ""),
                    "gene_biotype": str(r["gene_biotype"] or ""),
                    "type": "gene",
                }
                for r in rows
            ]
            return jsonify(features)

        rows = genome_lib.features_in_region(
            conn, chrom, start, end, ["transcript", "exon", "CDS"], 50000
        )
    finally:
        conn.close()

    transcripts: dict[str, dict[str, Any]] = {}
    for r in rows:
        tid = r["transcript_id"]
        if not tid:
            continue
        t = transcripts.get(tid)
        if t is None:
            gname = str(r["gene_name"] or "")
            # IGV groups features by `name`; if every isoform shares the gene
            # name, igv.js stacks them onto a single row, hiding alternative
            # splicing entirely. Use transcript_id as the unique track label
            # and keep gene_name in a separate field for popovers.
            t = {
                "chr": str(r["chrom"] or ""),
                "start": None,
                "end": None,
                "name": str(tid),
                "strand": str(r["strand"] or "."),
                "gene_id": str(r["gene_id"] or ""),
                "transcript_id": str(tid),
                "gene_name": gname,
                "gene_biotype": str(r["gene_biotype"] or ""),
                "transcript_biotype": str(r["transcript_biotype"] or ""),
                "biotype": str(r["transcript_biotype"] or r["gene_biotype"] or ""),
                "exons": [],
                "_cds": [],
            }
            transcripts[tid] = t
        rtype = r["type"]
        rs = int(r["start"]) - 1
        re_ = int(r["end"])
        if rtype == "transcript":
            t["start"] = rs
            t["end"] = re_
        elif rtype == "exon":
            t["exons"].append({"start": rs, "end": re_})
        elif rtype == "CDS":
            t["_cds"].append((rs, re_))

    out: list[dict[str, Any]] = []
    for t in transcripts.values():
        if not t["exons"] and t["start"] is None:
            continue
        if t["start"] is None or t["end"] is None:
            t["start"] = min(e["start"] for e in t["exons"]) if t["exons"] else 0
            t["end"] = max(e["end"] for e in t["exons"]) if t["exons"] else 0
        t["exons"].sort(key=lambda e: e["start"])
        cds = t.pop("_cds")
        if cds:
            cds.sort()
            cds_min = min(c[0] for c in cds)
            cds_max = max(c[1] for c in cds)
            for ex in t["exons"]:
                if ex["end"] <= cds_min or ex["start"] >= cds_max:
                    ex["cdStart"] = ex["start"]
                    ex["cdEnd"] = ex["start"]
                else:
                    ex["cdStart"] = max(ex["start"], cds_min)
                    ex["cdEnd"] = min(ex["end"], cds_max)
        out.append(t)

    out.sort(key=lambda x: (x["start"], x["end"]))
    return jsonify(out)


@app.route("/api/genome/igv.bed")
def api_genome_igv_bed():
    """Region-aware BED12 export consumed by the igv.js annotation track.

    Each visible transcript becomes one BED12 record (one row per isoform).
    `thickStart`/`thickEnd` mark the CDS extent so igv.js draws thick CDS
    boxes and thin UTR ends out-of-the-box. Blocks describe every exon, so
    introns are rendered as chevroned lines automatically.

    Why BED12 instead of a JSON custom source: igv.js's annotation track has
    rock-solid native handling for BED12 isoforms - in `EXPANDED` mode
    overlapping isoforms are guaranteed to land on separate rows. Custom
    JSON sources, in contrast, were silently collapsing all transcripts of a
    gene onto a single row in this build.
    """
    _require_genome()
    chrom, start, end = _parse_region_args()
    span = end - start + 1

    conn = _get_gtf_conn()
    try:
        if span > 5_000_000:
            rows = genome_lib.features_in_region(
                conn, chrom, start, end, ["gene"], MAX_FEATURES
            )
            lines: list[str] = []
            for r in rows:
                gname = str(r["gene_name"] or r["gene_id"] or "gene")
                cstart = max(0, int(r["start"]) - 1)
                cend = int(r["end"])
                strand = str(r["strand"] or ".")
                rgb = _BIOTYPE_RGB.get(str(r["gene_biotype"] or ""), "37,99,235")
                blen = cend - cstart
                lines.append(
                    "\t".join(
                        [
                            str(r["chrom"]),
                            str(cstart),
                            str(cend),
                            gname,
                            "0",
                            strand,
                            str(cstart),
                            str(cend),
                            rgb,
                            "1",
                            f"{blen},",
                            "0,",
                        ]
                    )
                )
            body = "\n".join(lines) + ("\n" if lines else "")
            return Response(body, mimetype="text/plain")

        rows = genome_lib.features_in_region(
            conn, chrom, start, end, ["transcript", "exon", "CDS"], 50000
        )
    finally:
        conn.close()

    transcripts: dict[str, dict[str, Any]] = {}
    for r in rows:
        tid = r["transcript_id"]
        if not tid:
            continue
        t = transcripts.get(tid)
        if t is None:
            t = {
                "chrom": str(r["chrom"] or ""),
                "name": str(tid),
                "strand": str(r["strand"] or "."),
                "gene_name": str(r["gene_name"] or ""),
                "biotype": str(r["transcript_biotype"] or r["gene_biotype"] or ""),
                "tx_start": None,
                "tx_end": None,
                "exons": [],
                "cds": [],
            }
            transcripts[tid] = t
        rs = int(r["start"]) - 1
        re_ = int(r["end"])
        rt = r["type"]
        if rt == "transcript":
            t["tx_start"] = rs
            t["tx_end"] = re_
        elif rt == "exon":
            t["exons"].append((rs, re_))
        elif rt == "CDS":
            t["cds"].append((rs, re_))

    out_lines: list[str] = []
    for t in transcripts.values():
        if not t["exons"] and t["tx_start"] is None:
            continue
        if t["tx_start"] is None or t["tx_end"] is None:
            t["tx_start"] = min(e[0] for e in t["exons"]) if t["exons"] else 0
            t["tx_end"] = max(e[1] for e in t["exons"]) if t["exons"] else 0
        cstart = t["tx_start"]
        cend = t["tx_end"]
        exons = sorted(t["exons"]) or [(cstart, cend)]
        if t["cds"]:
            thick_start = min(c[0] for c in t["cds"])
            thick_end = max(c[1] for c in t["cds"])
        else:
            thick_start = cend
            thick_end = cend
        block_count = len(exons)
        block_sizes = ",".join(str(e[1] - e[0]) for e in exons) + ","
        block_starts = ",".join(str(e[0] - cstart) for e in exons) + ","
        rgb = _BIOTYPE_RGB.get(t["biotype"], "37,99,235")
        # Use the gene name as a tertiary label so users still see the gene
        # symbol on the track via igv's "tracklabel" option, while keeping
        # transcript_id as the primary unique name (column 4).
        label = t["name"]
        if t["gene_name"] and t["gene_name"] != t["name"]:
            label = f"{t['name']} ({t['gene_name']})"
        out_lines.append(
            "\t".join(
                [
                    t["chrom"],
                    str(cstart),
                    str(cend),
                    label,
                    "0",
                    t["strand"],
                    str(thick_start),
                    str(thick_end),
                    rgb,
                    str(block_count),
                    block_sizes,
                    block_starts,
                ]
            )
        )

    out_lines.sort(key=lambda line: (
        int(line.split("\t", 3)[1]),
        int(line.split("\t", 3)[2]),
    ))
    body = "\n".join(out_lines) + ("\n" if out_lines else "")
    return Response(body, mimetype="text/plain")


@app.route("/api/genome/gene/<gene_id>")
def api_genome_gene(gene_id: str):
    """Return a gene plus all of its transcripts and their exon / CDS structure.

    Used by the right-side detail panel to show alternative-splicing isoforms.
    """
    _require_genome()
    conn = _get_gtf_conn()
    try:
        gene_row = conn.execute(
            "SELECT chrom, start, end, strand, gene_id, gene_name, gene_biotype "
            "FROM features WHERE type='gene' AND gene_id = ? LIMIT 1",
            (gene_id,),
        ).fetchone()
        if gene_row is None:
            abort(404, description=f"gene {gene_id} not found")
        rows = conn.execute(
            "SELECT id, type, start, end, strand, phase, transcript_id, gene_name, "
            "transcript_biotype, gene_biotype "
            "FROM features WHERE gene_id = ? AND type IN "
            "('transcript','exon','CDS','five_prime_utr','three_prime_utr',"
            "'start_codon','stop_codon') "
            "ORDER BY start, end",
            (gene_id,),
        ).fetchall()
    finally:
        conn.close()

    gene = {
        "gene_id": gene_row["gene_id"],
        "gene_name": gene_row["gene_name"] or gene_row["gene_id"],
        "gene_biotype": gene_row["gene_biotype"] or "",
        "chrom": gene_row["chrom"],
        "start": int(gene_row["start"]),
        "end": int(gene_row["end"]),
        "strand": gene_row["strand"] or ".",
        "length": int(gene_row["end"]) - int(gene_row["start"]) + 1,
    }

    by_tx: dict[str, dict[str, Any]] = {}
    for r in rows:
        tid = r["transcript_id"]
        if not tid:
            continue
        t = by_tx.get(tid)
        if t is None:
            t = {
                "transcript_id": tid,
                "transcript_biotype": r["transcript_biotype"] or "",
                "strand": r["strand"] or gene["strand"],
                "start": None,
                "end": None,
                "exons": [],
                "cds": [],
                "utrs": [],
                "start_codons": [],
                "stop_codons": [],
            }
            by_tx[tid] = t
        rt = r["type"]
        if rt == "transcript":
            t["start"] = int(r["start"])
            t["end"] = int(r["end"])
        elif rt == "exon":
            t["exons"].append({"start": int(r["start"]), "end": int(r["end"])})
        elif rt == "CDS":
            t["cds"].append({"start": int(r["start"]), "end": int(r["end"]), "phase": r["phase"]})
        elif rt in ("five_prime_utr", "three_prime_utr"):
            t["utrs"].append(
                {"type": rt, "start": int(r["start"]), "end": int(r["end"])}
            )
        elif rt == "start_codon":
            t["start_codons"].append({"start": int(r["start"]), "end": int(r["end"])})
        elif rt == "stop_codon":
            t["stop_codons"].append({"start": int(r["start"]), "end": int(r["end"])})

    transcripts = []
    for t in by_tx.values():
        if not t["exons"] and t["start"] is None:
            continue
        t["exons"].sort(key=lambda e: e["start"])
        if t["start"] is None or t["end"] is None:
            t["start"] = t["exons"][0]["start"] if t["exons"] else 0
            t["end"] = t["exons"][-1]["end"] if t["exons"] else 0
        t["cds"].sort(key=lambda c: c["start"])
        # Derived stats
        t["exon_count"] = len(t["exons"])
        t["length"] = sum(e["end"] - e["start"] + 1 for e in t["exons"])
        t["cds_length"] = sum(c["end"] - c["start"] + 1 for c in t["cds"])
        if t["cds"]:
            t["cds_min"] = t["cds"][0]["start"]
            t["cds_max"] = t["cds"][-1]["end"]
        else:
            t["cds_min"] = None
            t["cds_max"] = None
        transcripts.append(t)

    # Sort: protein_coding first, then by exon count desc, then by length desc.
    def _tx_sort_key(t):
        return (
            0 if t["transcript_biotype"] == "protein_coding" else 1,
            -t["exon_count"],
            -t["length"],
        )
    transcripts.sort(key=_tx_sort_key)
    return jsonify({"gene": gene, "transcripts": transcripts})


@app.route("/api/genome/transcript/<transcript_id>")
def api_genome_transcript(transcript_id: str):
    """Return a single transcript with full exon / CDS structure (no parent gene info)."""
    _require_genome()
    conn = _get_gtf_conn()
    try:
        rows = conn.execute(
            "SELECT type, chrom, start, end, strand, phase, gene_id, gene_name, "
            "gene_biotype, transcript_biotype "
            "FROM features WHERE transcript_id = ? AND type IN "
            "('transcript','exon','CDS','five_prime_utr','three_prime_utr',"
            "'start_codon','stop_codon') "
            "ORDER BY start, end",
            (transcript_id,),
        ).fetchall()
    finally:
        conn.close()
    if not rows:
        abort(404, description=f"transcript {transcript_id} not found")

    info = {
        "transcript_id": transcript_id,
        "gene_id": "",
        "gene_name": "",
        "gene_biotype": "",
        "transcript_biotype": "",
        "chrom": "",
        "strand": ".",
        "start": None,
        "end": None,
        "exons": [],
        "cds": [],
        "utrs": [],
        "start_codons": [],
        "stop_codons": [],
    }
    for r in rows:
        info["gene_id"] = info["gene_id"] or (r["gene_id"] or "")
        info["gene_name"] = info["gene_name"] or (r["gene_name"] or "")
        info["gene_biotype"] = info["gene_biotype"] or (r["gene_biotype"] or "")
        info["transcript_biotype"] = info["transcript_biotype"] or (r["transcript_biotype"] or "")
        info["chrom"] = info["chrom"] or (r["chrom"] or "")
        info["strand"] = r["strand"] or info["strand"]
        rt = r["type"]
        if rt == "transcript":
            info["start"] = int(r["start"])
            info["end"] = int(r["end"])
        elif rt == "exon":
            info["exons"].append({"start": int(r["start"]), "end": int(r["end"])})
        elif rt == "CDS":
            info["cds"].append({"start": int(r["start"]), "end": int(r["end"]), "phase": r["phase"]})
        elif rt in ("five_prime_utr", "three_prime_utr"):
            info["utrs"].append({"type": rt, "start": int(r["start"]), "end": int(r["end"])})
        elif rt == "start_codon":
            info["start_codons"].append({"start": int(r["start"]), "end": int(r["end"])})
        elif rt == "stop_codon":
            info["stop_codons"].append({"start": int(r["start"]), "end": int(r["end"])})
    info["exons"].sort(key=lambda e: e["start"])
    info["cds"].sort(key=lambda e: e["start"])
    if info["start"] is None and info["exons"]:
        info["start"] = info["exons"][0]["start"]
        info["end"] = info["exons"][-1]["end"]
    info["exon_count"] = len(info["exons"])
    info["length"] = sum(e["end"] - e["start"] + 1 for e in info["exons"])
    info["cds_length"] = sum(c["end"] - c["start"] + 1 for c in info["cds"])
    return jsonify(info)


@app.route("/api/genome/search")
def api_genome_search():
    _require_genome()
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"q": "", "count": 0, "items": []})
    try:
        limit = min(100, max(1, int(request.args.get("limit", "20"))))
    except ValueError:
        limit = 20
    conn = _get_gtf_conn()
    try:
        rows = genome_lib.search_features(conn, q, limit)
    finally:
        conn.close()
    items = [
        {
            "type": str(r.get("hit_type") or "gene"),
            "chrom": str(r["chrom"] or ""),
            "start": int(r["start"]),
            "end": int(r["end"]),
            "strand": str(r["strand"] or "."),
            "gene_id": str(r["gene_id"] or ""),
            "gene_name": str(r["gene_name"] or ""),
            "gene_biotype": str(r["gene_biotype"] or ""),
            "transcript_id": str(r.get("transcript_id") or "") if r.get("transcript_id") else "",
            "transcript_biotype": str(r.get("transcript_biotype") or "") if r.get("transcript_biotype") else "",
        }
        for r in rows
    ]
    return jsonify({"q": q, "count": len(items), "items": items})


@app.route("/api/genome/region/dna")
def api_genome_region_dna_export():
    """Return the region's DNA in FASTA format (60bp wrap, downloadable)."""
    _require_genome()
    chrom, start, end = _parse_region_args()
    span = end - start + 1
    if span > 50_000_000:
        abort(400, description="region too large for export (>50Mb)")
    seq = genome_lib.read_fasta_region(GENOME_FASTA, _get_fai(), chrom, start, end)
    body = f">{chrom}:{start}-{end}\n{fasta_wrap(seq)}\n"
    fname = f"{chrom}_{start}_{end}.fa"
    return Response(
        body,
        mimetype="text/x-fasta",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.route("/api/genome/region/gtf")
def api_genome_region_gtf_export():
    """Return GTF features overlapping the region as a downloadable .gtf file."""
    _require_genome()
    chrom, start, end = _parse_region_args()
    types_arg = request.args.get("types", "")
    types = [t.strip() for t in types_arg.split(",") if t.strip()] or None
    conn = _get_gtf_conn()
    try:
        rows = genome_lib.features_in_region(
            conn, chrom, start, end, types, limit=200000
        )
    finally:
        conn.close()
    out_lines: list[str] = [
        f"# Sus_scrofa.Sscrofa11.1.108 GTF export",
        f"# region: {chrom}:{start}-{end}",
        f"# features: {len(rows)}",
    ]
    for r in rows:
        attrs_parts = []
        if r["gene_id"]:
            attrs_parts.append(f'gene_id "{r["gene_id"]}"')
        if r["transcript_id"]:
            attrs_parts.append(f'transcript_id "{r["transcript_id"]}"')
        if r["gene_name"]:
            attrs_parts.append(f'gene_name "{r["gene_name"]}"')
        if r["gene_biotype"]:
            attrs_parts.append(f'gene_biotype "{r["gene_biotype"]}"')
        if r["transcript_biotype"]:
            attrs_parts.append(f'transcript_biotype "{r["transcript_biotype"]}"')
        attrs = "; ".join(attrs_parts) + (";" if attrs_parts else "")
        out_lines.append(
            "\t".join(
                [
                    r["chrom"],
                    r["source"] or ".",
                    r["type"],
                    str(r["start"]),
                    str(r["end"]),
                    ".",
                    r["strand"] or ".",
                    r["phase"] or ".",
                    attrs,
                ]
            )
        )
    body = "\n".join(out_lines) + "\n"
    fname = f"{chrom}_{start}_{end}.gtf"
    return Response(
        body,
        mimetype="text/plain",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# Static file endpoints required by igv.js (must support HTTP Range requests;
# Flask's send_file does this automatically).
@app.route("/genome/data/genome.fa")
def serve_genome_fasta():
    _require_genome()
    return send_file(
        GENOME_FASTA,
        mimetype="text/plain",
        conditional=True,
        download_name="Sus_scrofa.Sscrofa11.1.dna.toplevel.fa",
    )


@app.route("/genome/data/genome.fa.fai")
def serve_genome_fai():
    _require_genome()
    return send_file(GENOME_FAI, mimetype="text/plain", conditional=True)


@app.route("/genome/data/genome.bed")
def serve_genome_bed():
    """Whole-genome transcript BED12 consumed by the IGV transcripts track."""
    _require_genome()
    _ensure_genome_bed()
    return send_file(
        GENOME_BED,
        mimetype="text/plain",
        conditional=True,
        download_name="Sus_scrofa.Sscrofa11.1.108.bed",
    )


@app.route("/genome/data/genome.genes.bed")
def serve_genome_genes_bed():
    """Whole-genome gene-level BED9 for the IGV genes track."""
    _require_genome()
    _ensure_genome_bed()
    return send_file(
        GENOME_GENES_BED,
        mimetype="text/plain",
        conditional=True,
        download_name="Sus_scrofa.Sscrofa11.1.108.genes.bed",
    )


@app.route("/genome/data/perv.bed")
def serve_perv_bed():
    """PERV sequences BED6 for IGV annotation track."""
    _ensure_perv_bed()
    if not PERV_BED.exists():
        abort(404)
    return send_file(
        PERV_BED,
        mimetype="text/plain",
        conditional=True,
        download_name="perv_sequences.bed",
    )


@app.route("/api/genome/perv/list")
def api_perv_list():
    """Return all PERV sequences with transformed domain/ORF annotations."""
    seqs = _load_perv_list()
    return jsonify({"sequences": seqs, "total": len(seqs)})


@app.route("/genome/data/homologous_seq.bed")
def serve_homologous_seq_bed():
    """Homologous sequences BED6 for IGV annotation track."""
    _ensure_homologous_beds()
    if not HOMOLOGOUS_SEQ_BED.exists():
        abort(404)
    return send_file(
        HOMOLOGOUS_SEQ_BED,
        mimetype="text/plain",
        conditional=True,
        download_name="homologous_seq.bed",
    )


@app.route("/genome/data/homologous_locus.bed")
def serve_homologous_locus_bed():
    """Homologous loci BED6 for IGV annotation track."""
    _ensure_homologous_beds()
    if not HOMOLOGOUS_LOCUS_BED.exists():
        abort(404)
    return send_file(
        HOMOLOGOUS_LOCUS_BED,
        mimetype="text/plain",
        conditional=True,
        download_name="homologous_locus.bed",
    )


@app.route("/api/genome/homologous/list")
def api_homologous_list():
    """Return all 876 homologous sequences with full metadata."""
    seqs, _ = _load_homologous()
    return jsonify({"sequences": seqs, "total": len(seqs)})


@app.route("/api/genome/homologous/loci")
def api_homologous_loci():
    """Return 188 loci with aggregated species/group distributions."""
    _, loci_map = _load_homologous()
    def _locus_sort_key(locus: dict) -> int:
        lid = locus["locus_id"]
        try:
            return int(lid.split("_")[1]) if "_" in lid else 0
        except (ValueError, IndexError):
            return 0

    loci = sorted(loci_map.values(), key=_locus_sort_key)
    return jsonify({"loci": loci, "total": len(loci)})


# -- multi-omics bigwig -------------------------------------------------------

# Cache for multi-omics sample metadata {filename: {period, tissue, target, sample}}
_MULTIOMICS_META_CACHE: dict[str, dict] | None = None
_MULTIOMICS_INDEX_CACHE: list | None = None


def _load_multiomics_meta() -> dict[str, dict]:
    """Parse all.sample.info (TSV) into a filename → metadata dict.

    Columns (1-based): File.location, File.name, Period, Tissue, Sequence.target,
                       Replicates, Standardization.methods, File.type, Sample
    """
    global _MULTIOMICS_META_CACHE
    if _MULTIOMICS_META_CACHE is not None:
        return _MULTIOMICS_META_CACHE
    result: dict[str, dict] = {}
    if not MULTIOMICS_META.exists():
        _MULTIOMICS_META_CACHE = result
        return result
    try:
        with MULTIOMICS_META.open(encoding="utf-8") as fh:
            for i, line in enumerate(fh):
                if i == 0:
                    continue  # skip header
                parts = line.rstrip("\n").split("\t")
                if len(parts) < 9:
                    continue
                fname = parts[1].strip()
                if not fname:
                    continue
                rep = parts[5].strip() if len(parts) > 5 else ""
                std = parts[6].strip() if len(parts) > 6 else ""
                result[fname] = {
                    "period":       parts[2].strip(),
                    "tissue":       parts[3].strip(),
                    "target":       parts[4].strip(),
                    "replicates":   "" if rep == "." else rep,
                    "std_method":   "" if std == "." else std,
                    "sample":       parts[8].strip(),
                }
    except Exception as exc:
        print(f"[WARN] Could not load multiomics metadata: {exc}")
    _MULTIOMICS_META_CACHE = result
    return result


# Valid category names (data type folders) and their display order.
_MULTIOMICS_CATEGORIES = ["ATAC-seq", "ChIP-seq", "RNA-seq", "WGBS"]


_MULTIOMICS_INDEX_CACHE: list | None = None


@app.route("/api/multiomics/index")
def api_multiomics_index():
    """Return a tree of Multi-omics categories and their .bw files with metadata."""
    global _MULTIOMICS_INDEX_CACHE
    if _MULTIOMICS_INDEX_CACHE is not None:
        resp = jsonify({"categories": _MULTIOMICS_INDEX_CACHE})
        resp.headers["Cache-Control"] = "private, max-age=300"
        return resp

    meta = _load_multiomics_meta()
    categories = []
    if MULTIOMICS_DIR.is_dir():
        # Scan known categories first, then any remaining directories.
        seen: set[str] = set()
        cat_dirs: list[Path] = []
        for name in _MULTIOMICS_CATEGORIES:
            p = MULTIOMICS_DIR / name
            if p.is_dir():
                cat_dirs.append(p)
                seen.add(name)
        for p in sorted(MULTIOMICS_DIR.iterdir()):
            if p.is_dir() and p.name not in seen and not p.name.startswith("."):
                cat_dirs.append(p)

        for cat_dir in cat_dirs:
            bw_dir = cat_dir / "data_bw"
            if not bw_dir.is_dir():
                # Fallback: scan the category dir itself for .bw files
                bw_dir = cat_dir
            files = []
            for bw in sorted(bw_dir.glob("*.bw")):
                rel = f"/multiomics/data/{cat_dir.name}/{bw.name}"
                m = meta.get(bw.name, {})
                files.append({
                    "name":       bw.stem,
                    "filename":   bw.name,
                    "size":       bw.stat().st_size,
                    "url":        rel,
                    "period":     m.get("period",     ""),
                    "tissue":     m.get("tissue",     ""),
                    "target":     m.get("target",     ""),
                    "replicates": m.get("replicates", ""),
                    "std_method": m.get("std_method", ""),
                    "sample":     m.get("sample",     ""),
                })
            # Collect unique filter values for this category
            periods     = sorted({f["period"]     for f in files if f["period"]})
            tissues     = sorted({f["tissue"]     for f in files if f["tissue"]})
            targets     = sorted({f["target"]     for f in files if f["target"]})
            replicates  = sorted({f["replicates"] for f in files if f["replicates"]})
            std_methods = sorted({f["std_method"] for f in files if f["std_method"]})
            samples     = sorted({f["sample"]     for f in files if f["sample"]})
            categories.append({
                "id":    cat_dir.name,
                "label": cat_dir.name,
                "files": files,
                "filter_options": {
                    "periods":      periods,
                    "tissues":      tissues,
                    "targets":      targets,
                    "replicates":   replicates,
                    "std_methods":  std_methods,
                    "samples":      samples,
                },
            })
    _MULTIOMICS_INDEX_CACHE = categories
    resp = jsonify({"categories": categories})
    resp.headers["Cache-Control"] = "private, max-age=300"
    return resp


_MULTIOMICS_RECOMMENDED_CACHE: list | None = None


def _resolve_multiomics_bw(fname: str, path_hint: str = "") -> dict | None:
    """Resolve a .bw filename to an index-style file dict, or None if missing."""
    meta = _load_multiomics_meta()
    m = meta.get(fname, {})

    cat: str | None = None
    if path_hint:
        parts = Path(path_hint).parts
        for i, part in enumerate(parts):
            if part in ("new.Multi-omics", "Multi-omics") and i + 1 < len(parts):
                cat = parts[i + 1]
                break

    if not cat:
        for cat_name in _MULTIOMICS_CATEGORIES:
            p = MULTIOMICS_DIR / cat_name / "data_bw" / fname
            if p.is_file():
                cat = cat_name
                break
        if not cat and MULTIOMICS_DIR.is_dir():
            for p in sorted(MULTIOMICS_DIR.iterdir()):
                if p.is_dir() and not p.name.startswith("."):
                    candidate = p / "data_bw" / fname
                    if candidate.is_file():
                        cat = p.name
                        break

    if not cat:
        return None

    bw_path = MULTIOMICS_DIR / cat / "data_bw" / fname
    if not bw_path.is_file():
        bw_path = MULTIOMICS_DIR / cat / fname
    if not bw_path.is_file():
        return None

    return {
        "name":       bw_path.stem,
        "filename":   fname,
        "size":       bw_path.stat().st_size,
        "url":        f"/multiomics/data/{cat}/{fname}",
        "category":   cat,
        "period":     m.get("period",     ""),
        "tissue":     m.get("tissue",     ""),
        "target":     m.get("target",     ""),
        "replicates": m.get("replicates", ""),
        "std_method": m.get("std_method", ""),
        "sample":     m.get("sample",     ""),
    }


def _load_multiomics_recommended() -> list[dict]:
    """Parse represent.sample.info into grouped recommended track lists."""
    global _MULTIOMICS_RECOMMENDED_CACHE
    if _MULTIOMICS_RECOMMENDED_CACHE is not None:
        return _MULTIOMICS_RECOMMENDED_CACHE

    groups: dict[str, list[str]] = {}
    path_hints: dict[str, str] = {}
    if MULTIOMICS_REPRESENT.is_file():
        try:
            with MULTIOMICS_REPRESENT.open(encoding="utf-8") as fh:
                for line in fh:
                    parts = line.rstrip("\n").split("\t")
                    if len(parts) < 9:
                        continue
                    fname = parts[1].strip()
                    group = parts[8].strip()
                    if not fname or not group:
                        continue
                    groups.setdefault(group, []).append(fname)
                    path_hints[fname] = parts[0].strip()
        except Exception as exc:
            print(f"[WARN] Could not load recommended multiomics tracks: {exc}")

    result: list[dict] = []
    for group_name in sorted(groups.keys()):
        files: list[dict] = []
        seen: set[str] = set()
        for fname in groups[group_name]:
            if fname in seen:
                continue
            seen.add(fname)
            entry = _resolve_multiomics_bw(fname, path_hints.get(fname, ""))
            if entry:
                files.append(entry)
        if files:
            result.append({
                "id":    group_name,
                "label": group_name,
                "files": files,
            })

    _MULTIOMICS_RECOMMENDED_CACHE = result
    return result


@app.route("/api/multiomics/recommended")
def api_multiomics_recommended():
    """Return recommended multi-omics track groups from represent.sample.info."""
    return jsonify({"groups": _load_multiomics_recommended()})


# Assay → category mapping used by /api/multiomics/summary so the front-end
# knows which subfolder a .bw lives in (matches the file layout under
# new.Multi-omics/<category>/data_bw/).
_ASSAY_TO_CATEGORY = {
    "ATAC":     "ATAC-seq",
    "RNA":      "RNA-seq",
    "WGBS":     "WGBS",
    # everything else (histone marks, CTCF, Pol2) lives under ChIP-seq
}
# Tissues we render directly on the pig anatomy SVG (the rest are cell lines
# shown as side cards on the home page).
_HOMEPAGE_CELL_LINES = {"PIEC", "PK15", "ST"}


def _assay_to_category(assay: str) -> str:
    return _ASSAY_TO_CATEGORY.get(assay, "ChIP-seq")


@app.route("/api/multiomics/summary")
def api_multiomics_summary():
    """Aggregate the full multi-omics atlas for the home-page widget.

    Returns:
      - totals: total_files / samples / periods / tissues / assays
      - tissues: { <tissue>: {
            total_files, is_cell_line, periods, assays,
            matrix:        { <period>: { <assay>: count } },
            assay_summary: { <assay>: { count, category, filenames: [...] } },
        } }
      - real_tissues / cell_lines: ordered lists used by the front-end
    """
    meta = _load_multiomics_meta()

    tissues: dict[str, dict] = {}
    periods_all: set[str] = set()
    assays_all: set[str] = set()
    samples_all: set[str] = set()
    total_files = 0

    for fname, info in meta.items():
        tissue = info.get("tissue") or ""
        period = info.get("period") or ""
        assay  = info.get("target") or ""
        sample = info.get("sample") or ""
        if not (tissue and period and assay):
            continue
        total_files += 1
        periods_all.add(period)
        assays_all.add(assay)
        if sample:
            samples_all.add(sample)

        tdata = tissues.setdefault(tissue, {
            "total_files": 0,
            "is_cell_line": tissue in _HOMEPAGE_CELL_LINES,
            "periods": set(),
            "assays": set(),
            "matrix": {},
            "assay_summary": {},
            "period_summary": {},
            "all_filenames": [],
        })
        tdata["total_files"] += 1
        tdata["periods"].add(period)
        tdata["assays"].add(assay)
        tdata["all_filenames"].append(fname)

        row = tdata["matrix"].setdefault(period, {})
        row[assay] = row.get(assay, 0) + 1

        a_entry = tdata["assay_summary"].setdefault(assay, {
            "count":     0,
            "category":  _assay_to_category(assay),
            "filenames": [],
        })
        a_entry["count"] += 1
        a_entry["filenames"].append(fname)

        p_entry = tdata["period_summary"].setdefault(period, {
            "count":     0,
            "assays":    set(),
            "filenames": [],
        })
        p_entry["count"] += 1
        p_entry["assays"].add(assay)
        p_entry["filenames"].append(fname)

    # Sort periods using a sensible order (S < P21 < P50 < P100 < P180,
    # anything unexpected falls back to lexicographic at the tail).
    _period_rank = {"S": 0, "P21": 1, "P50": 2, "P100": 3, "P180": 4}
    def _period_key(p: str):
        return (_period_rank.get(p, 99), p)

    periods_sorted = sorted(periods_all, key=_period_key)
    assays_sorted  = sorted(assays_all)

    tissues_out: dict[str, dict] = {}
    for tname, td in tissues.items():
        # Materialise sets into sorted lists for JSON serialisation.
        period_summary_out = {}
        for p, pe in td["period_summary"].items():
            period_summary_out[p] = {
                "count":     pe["count"],
                "assays":    sorted(pe["assays"]),
                "filenames": pe["filenames"],
            }
        tissues_out[tname] = {
            "total_files":   td["total_files"],
            "is_cell_line":  td["is_cell_line"],
            "periods":       sorted(td["periods"], key=_period_key),
            "assays":        sorted(td["assays"]),
            "matrix":        td["matrix"],
            "assay_summary": td["assay_summary"],
            "period_summary": period_summary_out,
            "all_filenames": td["all_filenames"],
        }

    real_tissues = sorted(
        [t for t, td in tissues_out.items() if not td["is_cell_line"]]
    )
    cell_lines = sorted(
        [t for t, td in tissues_out.items() if td["is_cell_line"]]
    )

    return jsonify({
        "totals": {
            "total_files": total_files,
            "samples":     len(samples_all),
            "periods":     periods_sorted,
            "tissues":     real_tissues + cell_lines,
            "assays":      assays_sorted,
        },
        "tissues":      tissues_out,
        "real_tissues": real_tissues,
        "cell_lines":   cell_lines,
    })


@app.route("/multiomics/data/<category>/<path:filename>")
def serve_multiomics_bw(category: str, filename: str):
    """Serve a .bw file with HTTP Range support for igv.js streaming."""
    if not filename.endswith(".bw"):
        abort(404)
    # Prevent path traversal: reject embedded separators / dotdot
    if "/" in filename or "\\" in filename or ".." in filename:
        abort(400, description="invalid filename")
    cat_safe = category.replace("/", "").replace("..", "")
    if cat_safe != category:
        abort(400, description="invalid category")
    # Try new structure: {category}/data_bw/{filename}
    p = MULTIOMICS_DIR / category / "data_bw" / filename
    if not p.is_file():
        # Fallback: {category}/{filename}
        p = MULTIOMICS_DIR / category / filename
    if not p.is_file():
        abort(404)
    return send_file(p, mimetype="application/octet-stream", conditional=True)


# -- multi-omics visualization download ------------------------------------


@app.route("/api/download/resolve_region")
def api_download_resolve_region():
    """Resolve an annotation entity or custom coordinates to a genomic region.

    Query params
    ------------
    type : gene | transcript | perv | homo_seq | homo_locus | custom | position
    id   : entity identifier (gene/transcript/perv/homo_seq/homo_locus)
    chrom, start, end : integers (custom)
    chrom, pos, window : integers (position; window defaults to 10000)

    Returns {chrom, start, end, name, length}  (coordinates are 1-based).
    """
    rtype = request.args.get("type", "").strip()

    if rtype in ("gene", "transcript"):
        _require_genome()
        entity_id = request.args.get("id", "").strip()
        if not entity_id:
            abort(400, description="id is required")
        conn = _get_gtf_conn()
        try:
            rows = genome_lib.search_features(conn, entity_id, limit=1)
        finally:
            conn.close()
        if not rows:
            abort(404, description=f"No {rtype} found for: {entity_id!r}")
        r = rows[0]
        name = r.get("gene_name") or r.get("gene_id") or entity_id
        return jsonify({
            "chrom":  r["chrom"],
            "start":  r["start"],
            "end":    r["end"],
            "name":   name,
            "length": r["end"] - r["start"] + 1,
        })

    if rtype == "perv":
        entity_id = request.args.get("id", "").strip()
        if not entity_id:
            abort(400, description="id is required")
        seqs = _load_perv_list()
        match = next((s for s in seqs if s["name"] == entity_id), None)
        if not match:
            abort(404, description=f"PERV sequence not found: {entity_id!r}")
        return jsonify({
            "chrom":  match["chrom"],
            "start":  match["start"],
            "end":    match["end"],
            "name":   match["name"],
            "length": match["length"],
        })

    if rtype == "homo_seq":
        entity_id = request.args.get("id", "").strip()
        if not entity_id:
            abort(400, description="id is required")
        seqs, _ = _load_homologous()
        match = next((s for s in seqs if s["q_name"] == entity_id), None)
        if not match:
            abort(404, description=f"Homologous sequence not found: {entity_id!r}")
        if match["start"] is None or match["end"] is None:
            abort(404, description="Coordinates not available for this sequence")
        return jsonify({
            "chrom":  match["chrom"],
            "start":  match["start"],
            "end":    match["end"],
            "name":   match["q_name"],
            "length": match["end"] - match["start"] + 1,
        })

    if rtype == "homo_locus":
        entity_id = request.args.get("id", "").strip()
        if not entity_id:
            abort(400, description="id is required")
        _, loci_map = _load_homologous()
        match = loci_map.get(entity_id)
        if not match:
            abort(404, description=f"Homologous locus not found: {entity_id!r}")
        if match["start"] is None or match["end"] is None:
            abort(404, description="Coordinates not available for this locus")
        return jsonify({
            "chrom":  match["chrom"],
            "start":  match["start"],
            "end":    match["end"],
            "name":   entity_id,
            "length": match["end"] - match["start"] + 1,
        })

    if rtype == "custom":
        chrom = request.args.get("chrom", "").strip()
        if not chrom:
            abort(400, description="chrom is required")
        try:
            start = int(request.args.get("start", "0"))
            end   = int(request.args.get("end",   "0"))
        except ValueError:
            abort(400, description="start/end must be integers")
        if start < 1 or end < start:
            abort(400, description="start must be ≥ 1 and end ≥ start")
        return jsonify({
            "chrom":  chrom,
            "start":  start,
            "end":    end,
            "name":   f"{chrom}:{start:,}–{end:,}",
            "length": end - start + 1,
        })

    if rtype == "position":
        chrom = request.args.get("chrom", "").strip()
        if not chrom:
            abort(400, description="chrom is required")
        try:
            pos    = int(request.args.get("pos",    "0"))
            window = max(100, int(request.args.get("window", "10000")))
        except ValueError:
            abort(400, description="pos/window must be integers")
        if pos < 1:
            abort(400, description="pos must be ≥ 1")
        half  = window // 2
        start = max(1, pos - half)
        end   = pos + half
        return jsonify({
            "chrom":  chrom,
            "start":  start,
            "end":    end,
            "name":   f"{chrom}:{pos:,} ± {half:,} bp",
            "length": end - start + 1,
        })

    abort(400, description=(
        f"Unknown region type: {rtype!r}. "
        "Must be one of: gene, transcript, perv, homo_seq, homo_locus, custom, position"
    ))


@app.route("/api/download/generate", methods=["POST"])
def api_download_generate():
    """Generate multi-omics visualization files and return as download.

    Request body (JSON)
    -------------------
    chrom        : str
    start        : int  (1-based)
    end          : int  (1-based)
    upstream     : int  bp to extend upstream   (default 0)
    downstream   : int  bp to extend downstream (default 0)
    bw_tracks    : list[str]  relative paths "Category/file.bw"
    annot_tracks : list[str]  subset of [genes, transcripts, perv, homo_seq, homo_loci]
    format       : "pdf" | "svg" | "png"

    Returns a single file, or a ZIP when multiple bw_tracks are selected.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.ticker as mticker
    except ImportError:
        abort(503, description="matplotlib is not installed. Run: pip install matplotlib")

    try:
        import pyBigWig
    except ImportError:
        abort(503, description="pyBigWig is not installed. Run: pip install pyBigWig")

    body = request.get_json(force=True, silent=True) or {}

    chrom = str(body.get("chrom", "")).strip()
    if not chrom:
        abort(400, description="chrom is required")

    try:
        start = int(body.get("start", 0))
        end   = int(body.get("end",   0))
    except (TypeError, ValueError):
        abort(400, description="start/end must be integers")

    upstream   = max(0, int(body.get("upstream",   0) or 0))
    downstream = max(0, int(body.get("downstream", 0) or 0))

    plot_start = max(1, start - upstream)
    plot_end   = end + downstream
    span       = plot_end - plot_start + 1

    if span > 10_000_000:
        abort(400, description=(
            f"Region too large ({span:,} bp). Maximum is 10 Mb. "
            "Reduce the region or extension."
        ))
    if span <= 0:
        abort(400, description="Invalid region: end must be greater than start")

    bw_tracks    = [str(p) for p in (body.get("bw_tracks")    or [])]
    annot_tracks = [str(p) for p in (body.get("annot_tracks") or [])]
    fmt          = str(body.get("format", "pdf")).lower()
    if fmt not in ("pdf", "svg", "png"):
        fmt = "pdf"

    if not bw_tracks:
        abort(400, description="At least one BigWig track must be selected")

    mime_map = {
        "pdf": "application/pdf",
        "svg": "image/svg+xml",
        "png": "image/png",
    }

    results: list[tuple[str, bytes]] = []

    for rel_path in bw_tracks:
        if ".." in rel_path or rel_path.startswith("/"):
            continue
        parts      = rel_path.split("/")
        cat_name   = parts[0] if len(parts) > 1 else ""
        fname      = parts[-1]
        # Try new structure: {category}/data_bw/{filename}
        bw_path = MULTIOMICS_DIR / cat_name / "data_bw" / fname if cat_name else MULTIOMICS_DIR / rel_path
        if not bw_path.is_file():
            # Fallback: direct path
            bw_path = MULTIOMICS_DIR / rel_path
        if not bw_path.is_file() or not bw_path.name.endswith(".bw"):
            continue

        track_stem = bw_path.stem
        track_label = f"{cat_name}: {track_stem}" if cat_name else track_stem

        try:
            data = _render_bw_figure(
                bw_path=bw_path,
                track_label=track_label,
                chrom=chrom,
                start=plot_start,
                end=plot_end,
                annot_tracks=annot_tracks,
                fmt=fmt,
                plt=plt,
                pyBigWig=pyBigWig,
                mticker=mticker,
            )
        except Exception as exc:
            app.logger.error("Error rendering %s: %s", rel_path, exc, exc_info=True)
            continue

        fname = f"{track_stem}_{chrom}_{plot_start}_{plot_end}.{fmt}"
        results.append((fname, data))

    if not results:
        abort(500, description="Failed to render any tracks. Check server logs for details.")

    if len(results) == 1:
        fname, data = results[0]
        return Response(
            data,
            mimetype=mime_map[fmt],
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    # Multiple tracks → bundle into ZIP
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, data in results:
            zf.writestr(fname, data)
    buf.seek(0)
    zip_fname = f"multiomics_{chrom}_{plot_start}_{plot_end}.zip"
    return Response(
        buf.read(),
        mimetype="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{zip_fname}"'},
    )


# ── pyGenomeTracks async job pipeline ───────────────────────────────────────
#
# A separate rendering engine (pyGenomeTracks) is offered alongside the
# matplotlib `/api/download/generate` pathway. The matplotlib endpoint is
# kept unchanged for backwards compatibility. pyGenomeTracks runs as a
# subprocess and can take several seconds, so jobs are queued on a small
# thread pool and the front-end polls /api/pygt/status/<id> until done.

import concurrent.futures as _cf  # noqa: E402  (kept local to the feature)
import shutil as _shutil           # noqa: E402
import threading as _threading     # noqa: E402
import uuid as _uuid               # noqa: E402

from server import pygt_engine as _pygt  # noqa: E402

PYGT_JOBS_DIR = BASE_DIR / "tmp" / "pygt_jobs"
PYGT_JOBS_DIR.mkdir(parents=True, exist_ok=True)
PYGT_JOB_TTL_SECONDS = int(os.environ.get("PYGT_JOB_TTL_SECONDS", str(24 * 3600)))
PYGT_CLEANUP_INTERVAL_SECONDS = int(
    os.environ.get("PYGT_CLEANUP_INTERVAL_SECONDS", str(24 * 3600))
)
PYGT_CLEANUP_LOCK = PYGT_JOBS_DIR / ".cleanup.lock"

_pygt_executor = _cf.ThreadPoolExecutor(max_workers=2, thread_name_prefix="pygt")
_pygt_jobs_lock = _threading.Lock()
_pygt_jobs: dict[str, dict] = {}


def _pygt_cleanup_old_jobs() -> int:
    """Delete on-disk job directories older than PYGT_JOB_TTL_SECONDS.

    Only removes ``tmp/pygt_jobs/{job_id}/`` subdirectories; the parent
    ``tmp/pygt_jobs/`` folder is always kept.  Returns the number removed.
    """
    if not PYGT_JOBS_DIR.is_dir():
        return 0
    cutoff = time.time() - PYGT_JOB_TTL_SECONDS
    removed = 0
    for child in PYGT_JOBS_DIR.iterdir():
        try:
            if child.is_dir() and child.stat().st_mtime < cutoff:
                _shutil.rmtree(child, ignore_errors=True)
                removed += 1
        except OSError:
            pass
    return removed


def _pygt_try_cleanup_old_jobs() -> None:
    """Run cleanup if the inter-process lock can be acquired (gunicorn-safe)."""
    PYGT_JOBS_DIR.mkdir(parents=True, exist_ok=True)
    try:
        with open(PYGT_CLEANUP_LOCK, "w", encoding="utf-8") as lock_f:
            try:
                import fcntl

                fcntl.flock(lock_f.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            except (ImportError, BlockingIOError, OSError):
                return
            n = _pygt_cleanup_old_jobs()
            if n:
                app.logger.info(
                    "pyGenomeTracks cleanup: removed %d expired job(s) from %s",
                    n, PYGT_JOBS_DIR,
                )
    except OSError:
        pass


def _pygt_cleanup_loop() -> None:
    """Background loop: sweep expired job dirs every PYGT_CLEANUP_INTERVAL_SECONDS."""
    while True:
        try:
            _pygt_try_cleanup_old_jobs()
        except Exception:  # pragma: no cover - defensive
            app.logger.exception("pyGenomeTracks periodic cleanup failed")
        time.sleep(PYGT_CLEANUP_INTERVAL_SECONDS)


def _pygt_start_cleanup_thread() -> None:
    t = _threading.Thread(
        target=_pygt_cleanup_loop, name="pygt-cleanup", daemon=True,
    )
    t.start()


_pygt_try_cleanup_old_jobs()
_pygt_start_cleanup_thread()


# Status is persisted to disk as ``status.json`` inside each job directory so
# polling requests can be served by any gunicorn worker (workers don't share
# in-process memory). The in-memory dict is still used as a hot cache by the
# worker that owns the rendering thread.

def _pygt_status_path(job_id: str) -> Path:
    return PYGT_JOBS_DIR / job_id / "status.json"


def _pygt_write_status(job_id: str, data: dict) -> None:
    job_dir = PYGT_JOBS_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    tmp = job_dir / "status.json.tmp"
    tmp.write_text(json.dumps(data), encoding="utf-8")
    os.replace(tmp, _pygt_status_path(job_id))


def _pygt_read_status(job_id: str) -> dict | None:
    p = _pygt_status_path(job_id)
    if not p.is_file():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def _pygt_run_job(job_id: str, spec: _pygt.RenderSpec) -> None:
    """Worker task: render the figure and persist job state to disk."""
    job_dir = PYGT_JOBS_DIR / job_id
    base = {"fmt": spec.fmt, "region": spec.region_str, "submitted": time.time()}
    try:
        running = {**base, "state": "running"}
        _pygt_write_status(job_id, running)
        with _pygt_jobs_lock:
            _pygt_jobs[job_id] = running

        result = _pygt.render(spec, job_dir)

        done = {
            **base,
            "state": "done",
            "artifact": str(result.artifact_path),
            "warnings": result.warnings,
            "finished": time.time(),
        }
        _pygt_write_status(job_id, done)
        with _pygt_jobs_lock:
            _pygt_jobs[job_id] = done
    except _pygt.PygtError as exc:
        err = {**base, "state": "error", "error": str(exc), "finished": time.time()}
        _pygt_write_status(job_id, err)
        with _pygt_jobs_lock:
            _pygt_jobs[job_id] = err
    except Exception as exc:  # pragma: no cover - defensive
        app.logger.exception("pyGenomeTracks job %s crashed", job_id)
        err = {
            **base, "state": "error",
            "error": f"Internal error: {exc}", "finished": time.time(),
        }
        _pygt_write_status(job_id, err)
        with _pygt_jobs_lock:
            _pygt_jobs[job_id] = err


def _pygt_lookup_job(job_id: str) -> dict | None:
    """Return job metadata from in-memory cache, falling back to disk."""
    with _pygt_jobs_lock:
        job = _pygt_jobs.get(job_id)
    if job is not None:
        return job
    return _pygt_read_status(job_id)


@app.route("/api/pygt/categories")
def api_pygt_categories():
    """Return the allow-listed Multi-omics categories for the pyGenomeTracks UI."""
    return jsonify({"categories": list(_pygt.ALLOWED_CATEGORIES)})


@app.route("/api/pygt/submit", methods=["POST"])
def api_pygt_submit():
    """Queue a pyGenomeTracks rendering job. Returns ``{job_id}`` on success."""
    body = request.get_json(force=True, silent=True) or {}
    try:
        spec = _pygt.validate_spec(body)
    except _pygt.PygtError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        app.logger.exception("pyGenomeTracks submit validation crashed")
        return jsonify({"error": f"Invalid request: {exc}"}), 400

    job_id = _uuid.uuid4().hex[:16]
    queued = {
        "state": "queued",
        "submitted": time.time(),
        "fmt": spec.fmt,
        "region": spec.region_str,
    }
    _pygt_write_status(job_id, queued)
    with _pygt_jobs_lock:
        _pygt_jobs[job_id] = queued
    _pygt_executor.submit(_pygt_run_job, job_id, spec)
    return jsonify({"job_id": job_id})


@app.route("/api/pygt/status/<job_id>")
def api_pygt_status(job_id: str):
    """Return the current state of a job (queued/running/done/error)."""
    if not re.fullmatch(r"[0-9a-f]{8,32}", job_id):
        abort(400, description="invalid job_id")
    job = _pygt_lookup_job(job_id)
    if not job:
        return jsonify({"state": "unknown"}), 404
    return jsonify({
        "state":     job.get("state"),
        "error":     job.get("error", ""),
        "fmt":       job.get("fmt"),
        "region":    job.get("region"),
        "submitted": job.get("submitted"),
        "finished":  job.get("finished"),
        "warnings":  job.get("warnings") or [],
    })


@app.route("/api/pygt/result/<job_id>")
def api_pygt_result(job_id: str):
    """Fetch the rendered artifact, the generated ini, or the run log."""
    if not re.fullmatch(r"[0-9a-f]{8,32}", job_id):
        abort(400, description="invalid job_id")
    kind = (request.args.get("kind") or "image").lower()
    if kind not in ("image", "ini", "log"):
        abort(400, description="kind must be image|ini|log")

    job = _pygt_lookup_job(job_id)
    if not job:
        abort(404, description="unknown job_id")

    state = job.get("state")
    fmt = job.get("fmt", "pdf")
    if state == "error":
        abort(500, description=job.get("error", "job failed"))
    if state != "done":
        abort(409, description=f"job is {state}, not ready")

    job_dir = PYGT_JOBS_DIR / job_id
    if kind == "ini":
        ini_file = job_dir / "tracks.ini"
        if not ini_file.is_file():
            abort(410, description="ini has been cleaned up")
        return Response(ini_file.read_text(encoding="utf-8"),
                        mimetype="text/plain; charset=utf-8")
    if kind == "log":
        log_file = job_dir / "run.log"
        if not log_file.is_file():
            abort(410, description="log has been cleaned up")
        return Response(log_file.read_text(encoding="utf-8"),
                        mimetype="text/plain; charset=utf-8")

    artifact_path = Path(job.get("artifact") or (job_dir / f"out.{fmt}"))
    if not artifact_path.is_file():
        artifact_path = job_dir / f"out.{fmt}"
    if not artifact_path.is_file():
        abort(410, description="artifact has been cleaned up")
    mime_map = {"pdf": "application/pdf", "svg": "image/svg+xml", "png": "image/png"}
    fname = f"pygenometracks_{job_id}.{fmt}"
    return send_file(
        artifact_path, mimetype=mime_map.get(fmt, "application/octet-stream"),
        as_attachment=True, download_name=fname,
    )


def _parse_bed_region(
    bed_path: Path, chrom: str, start: int, end: int, *, is_bed12: bool = False
) -> list:
    """Parse BED features overlapping [start, end] (1-based coordinates).

    Returns list of (feat_start, feat_end, name, strand, blocks) where
    blocks is a list of (block_start, block_end) for BED12, else [].
    """
    features: list = []
    if not bed_path.exists():
        return features
    with open(bed_path) as fh:
        for line in fh:
            if line.startswith("#") or not line.strip():
                continue
            parts = line.rstrip("\n").split("\t")
            if len(parts) < 3 or parts[0].strip() != chrom:
                continue
            try:
                fs = int(parts[1]) + 1  # BED 0-based → 1-based
                fe = int(parts[2])
            except ValueError:
                continue
            if fe < start or fs > end:
                continue
            name   = parts[3].strip() if len(parts) > 3 else ""
            strand = parts[5].strip() if len(parts) > 5 else "+"

            blocks: list = []
            if is_bed12 and len(parts) >= 12:
                try:
                    blk_sizes  = [int(x) for x in parts[10].rstrip(",").split(",") if x]
                    blk_starts = [int(x) for x in parts[11].rstrip(",").split(",") if x]
                    for bs, sz in zip(blk_starts, blk_sizes):
                        blocks.append((fs + bs, fs + bs + sz))
                except (ValueError, IndexError):
                    blocks = []

            features.append((fs, fe, name, strand, blocks))
    return features


def _stack_features(features: list) -> tuple[list, int]:
    """Greedy interval scheduling: assign each feature the earliest non-overlapping row.

    Returns ([(row_idx, feat), ...], total_row_count).
    """
    sorted_feats = sorted(features, key=lambda f: f[0])
    row_ends: list[int] = []
    result: list = []
    for feat in sorted_feats:
        s = feat[0]
        placed = False
        for r, re in enumerate(row_ends):
            if s > re + 2:
                row_ends[r] = feat[1]
                result.append((r, feat))
                placed = True
                break
        if not placed:
            result.append((len(row_ends), feat))
            row_ends.append(feat[1])
    return result, max(1, len(row_ends))


def _render_bw_figure(
    *,
    bw_path: Path,
    track_label: str,
    chrom: str,
    start: int,
    end: int,
    annot_tracks: list[str],
    fmt: str,
    plt,
    pyBigWig,
    mticker,
) -> bytes:
    """Render one BigWig signal track + IGV-style expanded annotation tracks.

    Design goals (matching user feedback):
    - Width scales with region span → landscape for long regions.
    - Compact annotation track height (each feature row ≈ 0.30 in).
    - Individual feature stacking with ID labels (italic for genes).
    - BED12 exon blocks for Transcripts track.
    - Colors and track names match genome.js defaults.
    """

    # ── Colors / labels matching genome.js ─────────────────────────────────────
    ANNOT_COLOR = {
        "genes":       "#555555",   # dark grey (IGV default)
        "transcripts": "#b8860b",   # dark goldenrod
        "perv":        "#e05c2b",   # orange-red
        "homo_seq":    "#4a90e2",   # cornflower blue
        "homo_loci":   "#9b59b6",   # medium purple
    }
    ANNOT_LABEL = {
        "genes":       "Genes",
        "transcripts": "Transcripts",
        "perv":        "PERV",
        "homo_seq":    "Homologous Seq",
        "homo_loci":   "Homologous Loci",
    }
    ANNOT_BED = {
        "genes":       GENOME_GENES_BED,
        "transcripts": GENOME_BED,
        "perv":        PERV_BED,
        "homo_seq":    HOMOLOGOUS_SEQ_BED,
        "homo_loci":   HOMOLOGOUS_LOCUS_BED,
    }

    valid_annot = [t for t in annot_tracks if t in ANNOT_COLOR]
    span        = max(1, end - start)

    FONT        = 10.0    # base font size (pt)
    BW_H_IN     = 1.8     # BigWig subplot height (inches)
    ROW_H_IN    = 0.30    # height per stacked feature row (inches)
    MIN_TRACK   = 0.40    # minimum annotation track height (inches)

    # ── Parse BED features and compute row layouts ──────────────────────────────
    track_stacked: dict[str, list] = {}
    track_nrows:   dict[str, int]  = {}
    for tname in valid_annot:
        is_bed12 = (tname == "transcripts")
        feats = _parse_bed_region(ANNOT_BED[tname], chrom, start, end, is_bed12=is_bed12)
        stacked, n_rows = _stack_features(feats)
        track_stacked[tname] = stacked
        track_nrows[tname]   = n_rows

    # ── Figure geometry ─────────────────────────────────────────────────────────
    # Width: proportional to span (landscape for long regions, cap at 30 in)
    fig_w = min(30.0, max(14.0, 8.0 + span / 80_000.0))

    # Height: BW track + annotation track heights
    annot_h_list = [max(MIN_TRACK, track_nrows[t] * ROW_H_IN) for t in valid_annot]
    fig_h = 0.7 + BW_H_IN + sum(annot_h_list)

    height_ratios = [BW_H_IN] + annot_h_list
    n_subplots    = 1 + len(valid_annot)

    fig, axes_raw = plt.subplots(
        n_subplots, 1,
        figsize=(fig_w, fig_h),
        gridspec_kw={"height_ratios": height_ratios},
        sharex=True,
    )
    axes: list = [axes_raw] if n_subplots == 1 else list(axes_raw)

    for ax in axes:
        ax.set_xlim(start, end)

    # ── BigWig signal ───────────────────────────────────────────────────────────
    ax0 = axes[0]
    bw  = None
    try:
        bw = pyBigWig.open(str(bw_path))
        n_bins   = min(2000, max(200, span))
        raw_vals = bw.stats(chrom, start - 1, end, nBins=n_bins, type="mean") or []
        vals     = [v if v is not None else 0.0 for v in raw_vals]
        xs       = [start + span * i / n_bins for i in range(n_bins)]
        ax0.fill_between(xs, vals, color="#2563eb", alpha=0.80, linewidth=0)
        ax0.set_ylim(bottom=0)
        max_val = max(vals) if any(v > 0 for v in vals) else 1.0
        ax0.set_yticks([0, max_val])
        ax0.set_yticklabels(["0", f"{max_val:.2g}"], fontsize=FONT - 1)
    except Exception as exc:
        ax0.text(0.5, 0.5, f"BigWig error:\n{exc}",
                 transform=ax0.transAxes, ha="center", va="center",
                 color="#dc2626", fontsize=FONT)
    finally:
        if bw is not None:
            try: bw.close()
            except Exception: pass

    ax0.spines[["top", "right"]].set_visible(False)
    ax0.set_title(f"{track_label}  |  {chrom}:{start:,}–{end:,}",
                  fontsize=FONT + 3, fontweight="bold", pad=6)
    ax0.set_ylabel("Signal", fontsize=FONT, labelpad=4)
    ax0.tick_params(labelbottom=False, axis="x")

    # ── Annotation tracks (IGV expanded layout) ─────────────────────────────────
    for i, tname in enumerate(valid_annot):
        ax        = axes[i + 1]
        color     = ANNOT_COLOR[tname]
        label     = ANNOT_LABEL[tname]
        minus_col = _darken_hex(color, 0.68)
        stacked   = track_stacked[tname]
        n_rows    = track_nrows[tname]
        is_tx     = (tname == "transcripts")
        is_gene   = (tname == "genes")

        ax.set_ylim(-0.1, n_rows)
        ax.set_yticks([])
        ax.spines[["top", "right", "left"]].set_visible(False)
        ax.set_ylabel(label, fontsize=FONT, rotation=0, labelpad=4,
                      ha="right", va="center")

        BAR_H = 0.55  # fractional bar height within each row unit

        for row_idx, feat in stacked:
            fs, fe, name, strand, blocks = feat
            ds = max(fs, start)
            de = min(fe, end)
            if de <= ds:
                continue

            fc       = minus_col if strand == "-" else color
            y_center = row_idx + 0.5

            if is_tx and blocks:
                # Thin backbone spanning the full transcript
                ax.broken_barh(
                    [(ds, de - ds)], (y_center - 0.10, 0.20),
                    facecolors=fc, linewidth=0, alpha=0.45,
                )
                # Thick exon blocks
                for bstart, bend in blocks:
                    bds = max(bstart, start)
                    bde = min(bend, end)
                    if bde > bds:
                        ax.broken_barh(
                            [(bds, bde - bds)], (y_center - BAR_H / 2, BAR_H),
                            facecolors=fc, linewidth=0, alpha=0.85,
                        )
            else:
                ax.broken_barh(
                    [(ds, de - ds)], (y_center - BAR_H / 2, BAR_H),
                    facecolors=fc, linewidth=0, alpha=0.85,
                )

            # Feature ID label (italic for genes)
            if name:
                feat_frac = (min(fe, end) - max(fs, start)) / span
                if feat_frac > 0.015:
                    text_x   = (max(fs, start) + min(fe, end)) / 2
                    txt_size = max(FONT - 3.5, 6.5)
                    ax.text(
                        text_x, y_center, name,
                        ha="center", va="center",
                        fontsize=txt_size,
                        style="italic" if is_gene else "normal",
                        color="white",
                        clip_on=True,
                        fontweight="semibold",
                    )

        is_last = (i == len(valid_annot) - 1)
        ax.tick_params(labelbottom=is_last, labelsize=FONT - 1, axis="x")
        if is_last:
            ax.xaxis.set_major_formatter(
                mticker.FuncFormatter(lambda x, _: f"{int(x):,}")
            )
            ax.set_xlabel("Genomic position (bp)", fontsize=FONT)

    if not valid_annot:
        ax0.tick_params(labelbottom=True, labelsize=FONT - 1, axis="x")
        ax0.xaxis.set_major_formatter(
            mticker.FuncFormatter(lambda x, _: f"{int(x):,}")
        )
        ax0.set_xlabel("Genomic position (bp)", fontsize=FONT)

    fig.tight_layout(h_pad=0.15)

    buf = io.BytesIO()
    fig.savefig(buf, format=fmt, dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _darken_hex(hex_color: str, factor: float = 0.7) -> str:
    """Darken a #rrggbb hex colour by multiplying each channel by factor."""
    hc = hex_color.lstrip("#")
    r  = int(int(hc[0:2], 16) * factor)
    g  = int(int(hc[2:4], 16) * factor)
    b  = int(int(hc[4:6], 16) * factor)
    return f"#{r:02x}{g:02x}{b:02x}"


# -- downloads --------------------------------------------------------------

@app.route("/download/<path:filename>")
def download_file(filename: str):
    if filename not in DOWNLOAD_WHITELIST:
        abort(404)
    return send_from_directory(SEQ_DIR, filename, as_attachment=True)


# -- error handler ----------------------------------------------------------

@app.errorhandler(400)
@app.errorhandler(404)
@app.errorhandler(503)
def _err(e):
    return jsonify({"error": str(e.description) if hasattr(e, "description") else str(e)}), e.code


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "5000")), debug=True)
